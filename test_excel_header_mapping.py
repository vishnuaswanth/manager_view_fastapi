"""
Test Excel Header Mapping to verify correct field name mappings.

Tests that:
1. API field names map correctly to database columns
2. Database data maps correctly to Excel headers
3. Multi-level headers show correct field names in correct order
"""

import sys
import os

# Add project root to path
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    HistoryLogData,
    HistoryChangeRecord,
    generate_history_excel,
    _get_metric_display_name,
    CORE_FIELDS
)
from openpyxl import load_workbook


def test_field_name_mapping():
    """Test that API field names map correctly to display names."""
    print("\n" + "=" * 70)
    print("Test: Field Name Mapping")
    print("=" * 70)

    # API field names (used internally)
    api_fields = ['forecast', 'fte_req', 'fte_avail', 'capacity']

    # Expected display names (shown in Excel)
    expected_display = {
        'forecast': 'Client Forecast',
        'fte_req': 'FTE Required',
        'fte_avail': 'FTE Available',
        'capacity': 'Capacity'
    }

    print("\nAPI Field Name → Excel Display Name:")
    for api_field in api_fields:
        display_name = _get_metric_display_name(api_field)
        expected = expected_display[api_field]
        match = "✓" if display_name == expected else "✗"
        print(f"  {match} '{api_field}' → '{display_name}' (expected: '{expected}')")

        assert display_name == expected, f"Mismatch: {api_field} → {display_name} (expected {expected})"

    print("\n✓ All field name mappings are correct")


def test_core_fields_order():
    """Test that CORE_FIELDS constant matches expected order."""
    print("\n" + "=" * 70)
    print("Test: CORE_FIELDS Order")
    print("=" * 70)

    expected_order = ["Client Forecast", "FTE Required", "FTE Available", "Capacity"]

    print("\nCORE_FIELDS constant:")
    for i, field in enumerate(CORE_FIELDS):
        expected = expected_order[i]
        match = "✓" if field == expected else "✗"
        print(f"  {match} Position {i}: '{field}' (expected: '{expected}')")

    assert CORE_FIELDS == expected_order, f"Order mismatch: {CORE_FIELDS} != {expected_order}"

    print("\n✓ CORE_FIELDS order is correct")


def test_excel_column_headers():
    """Test actual Excel output has correct headers in correct order."""
    print("\n" + "=" * 70)
    print("Test: Excel Column Headers")
    print("=" * 70)

    # Create sample history data with all 4 fields for one month
    history_log_data = HistoryLogData(
        id="test-123",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T10:30:00Z",
        user="test_user",
        description="Test header mapping",
        records_modified=1,
        summary_data=None
    )

    # Create changes for all 4 core fields
    changes = []
    month_label = "Jun-25"
    api_fields = ['forecast', 'fte_req', 'fte_avail', 'capacity']
    test_values = {
        'forecast': (1000, 1000, 0),     # old, new, delta
        'fte_req': (20, 20, 0),
        'fte_avail': (20, 25, 5),
        'capacity': (1000, 1125, 125)
    }

    for api_field in api_fields:
        old_val, new_val, delta = test_values[api_field]
        changes.append(HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name=f"{month_label}.{api_field}",
            old_value=old_val,
            new_value=new_val,
            delta=delta,
            month_label=month_label
        ))

    # Generate Excel
    excel_buffer = generate_history_excel(history_log_data, changes)

    # Load and verify headers
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print("\nRow 2 (Field Headers under Jun-25):")
    # Find the Jun-25 column (should be column 6 if there are 5 static columns)
    month_col_start = 6  # After Main LOB, State, Case Type, Case ID, Target CPH

    # Get field headers from row 2
    actual_headers = []
    for col_idx in range(month_col_start, month_col_start + 4):
        cell = ws.cell(row=2, column=col_idx)
        actual_headers.append(cell.value)
        print(f"  Column {col_idx}: {cell.value}")

    # Verify against CORE_FIELDS
    print("\nVerifying against CORE_FIELDS:")
    for i, (expected, actual) in enumerate(zip(CORE_FIELDS, actual_headers)):
        match = "✓" if expected == actual else "✗"
        print(f"  {match} Position {i}: Expected '{expected}', Got '{actual}'")

        if expected != actual:
            print(f"\n✗ MISMATCH at position {i}:")
            print(f"  Expected: '{expected}'")
            print(f"  Got:      '{actual}'")
            raise AssertionError(f"Header mismatch at position {i}: {actual} != {expected}")

    print("\n✓ All Excel headers match CORE_FIELDS in correct order")


def test_bench_allocation_field_values():
    """Test that bench allocation shows correct values for each field."""
    print("\n" + "=" * 70)
    print("Test: Bench Allocation Field Values")
    print("=" * 70)

    print("\nIn Bench Allocation:")
    print("  - Client Forecast: NEVER changes (forecast_change=0)")
    print("  - FTE Required: NEVER changes (fte_req_change=0)")
    print("  - FTE Available: CHANGES (this is what we allocate)")
    print("  - Capacity: CHANGES (recalculated based on FTE Available)")

    # Create sample data representing bench allocation
    history_log_data = HistoryLogData(
        id="test-456",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T10:30:00Z",
        user="system",
        description="Bench allocation test",
        records_modified=1,
        summary_data=None
    )

    # Bench allocation: only fte_avail and capacity change
    changes = [
        # Forecast - unchanged (old=new)
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.forecast",
            old_value=1000,
            new_value=1000,  # Same as old
            delta=0,
            month_label="Jun-25"
        ),
        # FTE Required - unchanged (old=new)
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_req",
            old_value=20,
            new_value=20,  # Same as old
            delta=0,
            month_label="Jun-25"
        ),
        # FTE Available - CHANGED
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,  # Changed from 20 to 25
            delta=5,
            month_label="Jun-25"
        ),
        # Capacity - CHANGED
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1125,  # Changed
            delta=125,
            month_label="Jun-25"
        )
    ]

    # Generate Excel
    excel_buffer = generate_history_excel(history_log_data, changes)
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    # Get data row (row 3)
    print("\nData Row (Row 3) - Values:")
    # Assuming Jun-25 starts at column 6
    col_start = 6
    field_names = ["Client Forecast", "FTE Required", "FTE Available", "Capacity"]

    for i, field_name in enumerate(field_names):
        col_idx = col_start + i
        cell_value = ws.cell(row=3, column=col_idx).value
        print(f"  {field_name}: {cell_value}")

    # Verify unchanged fields show as single value (1000, 20)
    # Verify changed fields show as "new (old)" format
    forecast_cell = ws.cell(row=3, column=col_start).value
    fte_req_cell = ws.cell(row=3, column=col_start + 1).value
    fte_avail_cell = ws.cell(row=3, column=col_start + 2).value
    capacity_cell = ws.cell(row=3, column=col_start + 3).value

    print("\nVerifying values:")
    print(f"  Client Forecast (unchanged): {forecast_cell}")
    assert forecast_cell == 1000 or forecast_cell == "1000", f"Expected 1000, got {forecast_cell}"

    print(f"  FTE Required (unchanged): {fte_req_cell}")
    assert fte_req_cell == 20 or fte_req_cell == "20", f"Expected 20, got {fte_req_cell}"

    print(f"  FTE Available (changed): {fte_avail_cell}")
    assert "25" in str(fte_avail_cell) and "20" in str(fte_avail_cell), \
        f"Expected '25 (20)', got {fte_avail_cell}"

    print(f"  Capacity (changed): {capacity_cell}")
    assert "1125" in str(capacity_cell) and "1000" in str(capacity_cell), \
        f"Expected '1125 (1000)', got {capacity_cell}"

    print("\n✓ All field values are correct (unchanged fields show single value, changed fields show new (old))")


def main():
    """Run all mapping tests."""
    print("\n" + "=" * 70)
    print("EXCEL HEADER MAPPING TEST SUITE")
    print("=" * 70)

    try:
        test_field_name_mapping()
        test_core_fields_order()
        test_excel_column_headers()
        test_bench_allocation_field_values()

        print("\n" + "=" * 70)
        print("✓ ALL MAPPING TESTS PASSED")
        print("=" * 70)
        print("\nField Name Mapping Verified:")
        print("  ✓ API field names ('forecast', 'fte_req', etc.) map correctly")
        print("  ✓ Display names ('Client Forecast', 'FTE Required', etc.) are correct")
        print("  ✓ CORE_FIELDS order matches expected order")
        print("  ✓ Excel headers appear in correct order")
        print("  ✓ Bench allocation values are correct (unchanged fields, changed fields)")

        return 0

    except AssertionError as e:
        print("\n" + "=" * 70)
        print("✗ TEST FAILED")
        print("=" * 70)
        print(f"\nAssertion Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    except Exception as e:
        print("\n" + "=" * 70)
        print("✗ UNEXPECTED ERROR")
        print("=" * 70)
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())
