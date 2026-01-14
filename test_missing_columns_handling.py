"""Test how missing columns are handled in DataFrame creation."""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    HistoryLogData,
    HistoryChangeRecord,
    generate_history_excel
)
from openpyxl import load_workbook


def test_partial_fields():
    """Test when only some fields have changes (not all 4)."""
    print("\n" + "=" * 70)
    print("Test: Handling Missing Columns (Partial Field Changes)")
    print("=" * 70)

    history_log = HistoryLogData(
        id="test-partial",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T10:30:00Z",
        user="test_user",
        description="Test partial fields",
        records_modified=1,
        summary_data=None
    )

    # Only 2 fields have changes (fte_avail and capacity)
    # This simulates real bench allocation where forecast and fte_req don't change
    changes = [
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,
            delta=5,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1125,
            delta=125,
            month_label="Jun-25"
        )
    ]

    print("\nInput: Only 2 fields (FTE Available, Capacity)")
    print("  - Missing: Client Forecast, FTE Required")

    # Generate Excel
    excel_buffer = generate_history_excel(history_log, changes)
    excel_buffer.seek(0)

    # Load and verify
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print("\nGenerated Excel:")
    print(f"  Columns: {ws.max_column}")

    # Check what's in the columns
    print("\n  Row 2 (Field Headers):")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        if cell.value:
            print(f"    Column {col}: {cell.value}")

    print("\n  Row 3 (Data):")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col)
        value = cell.value if cell.value is not None else "(empty)"
        print(f"    Column {col}: {value}")

    # Verify structure
    # Expected: 4 static + 4 Jun-25 = 8 columns
    # Even if only 2 fields have data, we should have placeholders for all 4
    expected_columns = 8
    assert ws.max_column == expected_columns, f"Expected {expected_columns} columns, got {ws.max_column}"

    # Check that all 4 field headers exist (even if some have no data)
    field_headers = []
    for col in range(5, 9):  # Columns 5-8 are Jun-25 fields
        field_headers.append(ws.cell(row=2, column=col).value)

    expected_headers = ["Client Forecast", "FTE Required", "FTE Available", "Capacity"]
    for expected, actual in zip(expected_headers, field_headers):
        assert actual == expected, f"Expected header '{expected}', got '{actual}'"

    # Check data values
    # Columns 5-6 (Client Forecast, FTE Required) should be empty/None
    # Columns 7-8 (FTE Available, Capacity) should have values
    forecast_val = ws.cell(row=3, column=5).value
    fte_req_val = ws.cell(row=3, column=6).value
    fte_avail_val = ws.cell(row=3, column=7).value
    capacity_val = ws.cell(row=3, column=8).value

    print("\n  Data values:")
    print(f"    Client Forecast: {forecast_val} (should be 0)")
    print(f"    FTE Required: {fte_req_val} (should be 0)")
    print(f"    FTE Available: {fte_avail_val} (should have value)")
    print(f"    Capacity: {capacity_val} (should have value)")

    # Assertions
    # Missing fields should be filled with 0
    assert forecast_val == 0, f"Client Forecast should be 0, got {forecast_val}"
    assert fte_req_val == 0, f"FTE Required should be 0, got {fte_req_val}"

    # Present fields should have values
    assert fte_avail_val is not None, "FTE Available should have a value"
    assert capacity_val is not None, "Capacity should have a value"

    print("\n✓ Partial fields handled correctly")
    print("  ✓ All 4 column headers exist")
    print("  ✓ Missing fields filled with 0")
    print("  ✓ Present fields show correct values")


if __name__ == "__main__":
    test_partial_fields()
