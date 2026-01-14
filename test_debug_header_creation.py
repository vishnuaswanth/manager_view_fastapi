"""Debug multi-level header creation to find why Client Forecast is missing."""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    HistoryLogData,
    HistoryChangeRecord,
    generate_history_excel,
    _prepare_pivot_data
)
from openpyxl import load_workbook
import pandas as pd
from io import BytesIO


def test_excel_before_and_after_headers():
    """Check Excel structure before and after multi-level headers."""
    print("\n" + "=" * 70)
    print("Debug: Excel Structure Before/After Multi-Level Headers")
    print("=" * 70)

    # Create sample data
    changes = [
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.forecast",
            old_value=1000,
            new_value=1000,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_req",
            old_value=20,
            new_value=20,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,
            delta=5,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1125,
            delta=125,
            month_label="Jun-25"
        )
    ]

    # Get pivot data
    pivot_data, month_labels, static_columns = _prepare_pivot_data(changes)

    print("\n1. Pivot Data Structure:")
    print(f"   Rows: {len(pivot_data)}")
    if pivot_data:
        print(f"   Columns: {list(pivot_data[0].keys())}")

    # Create DataFrame (what pandas sees)
    df = pd.DataFrame(pivot_data)

    print("\n2. DataFrame Columns (before Excel):")
    for i, col in enumerate(df.columns):
        print(f"   {i}: {col}")

    # Write to Excel WITHOUT multi-level headers (just pandas default)
    excel_buffer_simple = BytesIO()
    with pd.ExcelWriter(excel_buffer_simple, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Changes', index=False)

    excel_buffer_simple.seek(0)
    wb_simple = load_workbook(excel_buffer_simple)
    ws_simple = wb_simple['Changes']

    print("\n3. Excel Columns (pandas default, row 1):")
    for col_idx in range(1, ws_simple.max_column + 1):
        cell = ws_simple.cell(row=1, column=col_idx)
        print(f"   Column {col_idx}: {cell.value}")

    print(f"\n   Total columns in simple Excel: {ws_simple.max_column}")

    # Now test with multi-level headers (full function)
    history_log_data = HistoryLogData(
        id="test-123",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T10:30:00Z",
        user="test_user",
        description="Test",
        records_modified=1,
        summary_data=None
    )

    excel_buffer_full = generate_history_excel(history_log_data, changes)
    excel_buffer_full.seek(0)
    wb_full = load_workbook(excel_buffer_full)
    ws_full = wb_full['Changes']

    print("\n4. Excel After Multi-Level Headers:")
    print(f"   Total columns: {ws_full.max_column}")
    print(f"   Total rows: {ws_full.max_row}")

    print("\n   Row 1 (Month Headers):")
    for col_idx in range(1, min(ws_full.max_column + 1, 15)):
        cell = ws_full.cell(row=1, column=col_idx)
        if cell.value:
            print(f"     Column {col_idx}: {cell.value}")

    print("\n   Row 2 (Field Headers):")
    for col_idx in range(1, min(ws_full.max_column + 1, 15)):
        cell = ws_full.cell(row=2, column=col_idx)
        if cell.value:
            print(f"     Column {col_idx}: {cell.value}")

    print("\n   Row 3 (Data):")
    for col_idx in range(1, min(ws_full.max_column + 1, 15)):
        cell = ws_full.cell(row=3, column=col_idx)
        if cell.value:
            print(f"     Column {col_idx}: {cell.value}")

    # Find where Jun-25 data should start
    print("\n5. Looking for Jun-25 columns:")
    jun_25_cols = []
    for col_idx in range(1, ws_full.max_column + 1):
        # Check row 1 for "Jun-25"
        cell_r1 = ws_full.cell(row=1, column=col_idx)
        # Check row 2 for field names
        cell_r2 = ws_full.cell(row=2, column=col_idx)
        # Check row 3 for data
        cell_r3 = ws_full.cell(row=3, column=col_idx)

        if cell_r1.value == "Jun-25" or (cell_r2.value and "Jun-25" not in str(cell_r2.value) and cell_r2.value in ["Client Forecast", "FTE Required", "FTE Available", "Capacity"]):
            jun_25_cols.append({
                'col': col_idx,
                'month_header': cell_r1.value,
                'field_header': cell_r2.value,
                'data': cell_r3.value
            })

    if jun_25_cols:
        print(f"   Found {len(jun_25_cols)} Jun-25 columns:")
        for col_info in jun_25_cols:
            print(f"     Col {col_info['col']}: {col_info['field_header']} = {col_info['data']}")
    else:
        print("   ✗ No Jun-25 columns found!")


if __name__ == "__main__":
    test_excel_before_and_after_headers()
