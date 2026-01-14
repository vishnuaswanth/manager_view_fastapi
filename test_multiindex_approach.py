"""
Test using pandas MultiIndex for multi-level headers.
This is the "proper" pandas way to create multi-level column headers.
"""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

import pandas as pd
from io import BytesIO
import tempfile
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_excel_with_multiindex():
    """Create Excel with MultiIndex columns - the pandas-native approach."""
    print("\n" + "=" * 70)
    print("TESTING MULTIINDEX APPROACH")
    print("=" * 70)

    # Sample data
    data = {
        'Amisys Medicaid DOMESTIC, TX, Claims, CL-001': {
            'Main LOB': 'Amisys Medicaid DOMESTIC',
            'State': 'TX',
            'Case Type': 'Claims',
            'Case ID': 'CL-001',
            'Jun-25_forecast': '1000',
            'Jun-25_fte_req': '20',
            'Jun-25_fte_avail': '25 (20)',
            'Jun-25_capacity': '1125 (1000)',
        },
        'Facets Commercial DOMESTIC, CA, Enrollment, EN-002': {
            'Main LOB': 'Facets Commercial DOMESTIC',
            'State': 'CA',
            'Case Type': 'Enrollment',
            'Case ID': 'EN-002',
            'Jul-25_forecast': '2000',
            'Jul-25_fte_req': '40',
            'Jul-25_fte_avail': '50 (40)',
            'Jul-25_capacity': '2500 (2000)',
        }
    }

    # Convert to DataFrame
    df = pd.DataFrame.from_dict(data, orient='index')

    # Create MultiIndex columns
    # Level 0: Month headers (or static column name for static cols)
    # Level 1: Field headers (or empty for static cols that merge vertically)

    columns_level_0 = []  # Month headers
    columns_level_1 = []  # Field headers

    for col in df.columns:
        if col in ['Main LOB', 'State', 'Case Type', 'Case ID']:
            # Static columns: same value in both levels (will be merged vertically in Excel)
            columns_level_0.append(col)
            columns_level_1.append('')  # Empty string for level 1
        else:
            # Dynamic columns: split "Jun-25_forecast" into ("Jun-25", "Client Forecast")
            parts = col.split('_', 1)
            if len(parts) == 2:
                month, field = parts
                field_display = {
                    'forecast': 'Client Forecast',
                    'fte_req': 'FTE Required',
                    'fte_avail': 'FTE Available',
                    'capacity': 'Capacity'
                }.get(field, field)

                columns_level_0.append(month)
                columns_level_1.append(field_display)
            else:
                columns_level_0.append(col)
                columns_level_1.append('')

    # Set MultiIndex
    df.columns = pd.MultiIndex.from_arrays([columns_level_0, columns_level_1])

    print("\n1. DataFrame with MultiIndex:")
    print(df)

    # Write to Excel
    print("\n2. Writing to Excel...")
    excel_buffer = BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Changes', index=False)

    excel_buffer.seek(0)

    # Load and inspect
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print(f"\n3. Excel structure:")
    print(f"   Rows: {ws.max_row}")
    print(f"   Columns: {ws.max_column}")

    print("\n   Row 1:")
    for col_idx in range(1, min(ws.max_column + 1, 15)):
        cell = ws.cell(row=1, column=col_idx)
        print(f"     Col {col_idx}: '{cell.value}'")

    print("\n   Row 2:")
    for col_idx in range(1, min(ws.max_column + 1, 15)):
        cell = ws.cell(row=2, column=col_idx)
        print(f"     Col {col_idx}: '{cell.value}'")

    # Check merged cells
    merged_ranges = list(ws.merged_cells.ranges)
    print(f"\n   Merged cell ranges: {len(merged_ranges)}")
    for mr in merged_ranges:
        print(f"     - {mr}")

    # Apply styling
    print("\n4. Applying styling...")

    header_font = Font(bold=True, size=11, color="FFFFFF")
    month_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    field_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Style both header rows
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = month_header_fill
        cell.alignment = header_alignment
        cell.border = border

    for cell in ws[2]:
        cell.font = header_font
        cell.fill = field_header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Style data rows
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = border

    # Save
    excel_buffer_styled = BytesIO()
    wb.save(excel_buffer_styled)

    # Save to file
    excel_buffer_styled.seek(0)
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        tmp.write(excel_buffer_styled.read())
        tmp_path = tmp.name

    print(f"\n5. Saved to: {tmp_path}")

    print("\n" + "=" * 70)
    print("✓ MULTIINDEX TEST COMPLETE")
    print("=" * 70)
    print("\nPlease open this file and check if Excel requires repair.")

    return tmp_path


if __name__ == "__main__":
    try:
        path = create_excel_with_multiindex()
        print(f"\n✓ File: {path}")
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
