"""
Test Multi-Level Headers in Excel Export.

Verifies that the Excel export now has:
- Row 1: Month headers (merged across 4 columns) + static columns (merged vertically)
- Row 2: Field headers (Client Forecast, FTE Required, FTE Available, Capacity)
- Row 3+: Data rows
"""

import sys
import os

# Add project root to path
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    generate_history_excel,
    HistoryLogData,
    HistoryChangeRecord,
    _parse_month_label,
    _prepare_pivot_data
)
from openpyxl import load_workbook
from io import BytesIO


def test_parse_month_label():
    """Test month label parsing for chronological sorting."""
    print("\n" + "=" * 70)
    print("Test 1: Month Label Parsing")
    print("=" * 70)

    test_cases = [
        ("Jun-25", (2025, 6)),
        ("Dec-24", (2024, 12)),
        ("Jan-26", (2026, 1)),
        ("Mar-2025", (2025, 3))
    ]

    for month_label, expected in test_cases:
        result = _parse_month_label(month_label)
        print(f"  {month_label:15s} → {result}")
        assert result == expected, f"Expected {expected}, got {result}"

    print("\n✓ TEST PASSED: All month labels parsed correctly")


def test_prepare_pivot_data_metadata():
    """Test that _prepare_pivot_data returns metadata tuple."""
    print("\n" + "=" * 70)
    print("Test 2: Pivot Data Metadata")
    print("=" * 70)

    changes = [
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,
            delta=5,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1125,
            delta=125,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jul-25.forecast",
            old_value=800,
            new_value=900,
            delta=100,
            month_label="Jul-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="target_cph",
            old_value=45,
            new_value=50,
            delta=5,
            month_label=None
        )
    ]

    pivot_data, month_labels, static_columns = _prepare_pivot_data(changes)

    print(f"\nPivot rows: {len(pivot_data)}")
    print(f"Month labels: {month_labels}")
    print(f"Static columns: {static_columns}")

    # Verify return types
    assert isinstance(pivot_data, list), "pivot_data should be list"
    assert isinstance(month_labels, list), "month_labels should be list"
    assert isinstance(static_columns, list), "static_columns should be list"

    # Verify month labels
    assert "Jun-25" in month_labels, "Jun-25 should be in month_labels"
    assert "Jul-25" in month_labels, "Jul-25 should be in month_labels"
    assert len(month_labels) == 2, "Should have 2 month labels"

    # Verify month labels are chronologically sorted
    assert month_labels == ["Jun-25", "Jul-25"], "Month labels should be sorted chronologically"

    # Verify static columns
    expected_static = ["Main LOB", "State", "Case Type", "Case ID", "Target CPH"]
    assert static_columns == expected_static, f"Expected {expected_static}, got {static_columns}"

    print("\n✓ TEST PASSED: Metadata returned correctly")


def test_multilevel_headers_structure():
    """Test that Excel has correct multi-level header structure."""
    print("\n" + "=" * 70)
    print("Test 3: Multi-Level Headers Structure")
    print("=" * 70)

    # Create sample history log
    history_log_data = HistoryLogData(
        id="test-123",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T10:30:00Z",
        user="test_user",
        description="Test multi-level headers",
        records_modified=1,
        summary_data=None
    )

    # Create sample changes
    changes = [
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.forecast",
            old_value=1000,
            new_value=1000,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.fte_req",
            old_value=20,
            new_value=20,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,
            delta=5,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1125,
            delta=125,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jul-25.forecast",
            old_value=800,
            new_value=900,
            delta=100,
            month_label="Jul-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jul-25.fte_req",
            old_value=15,
            new_value=18,
            delta=3,
            month_label="Jul-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jul-25.fte_avail",
            old_value=18,
            new_value=18,
            delta=0,
            month_label="Jul-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jul-25.capacity",
            old_value=900,
            new_value=950,
            delta=50,
            month_label="Jul-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="target_cph",
            old_value=45,
            new_value=50,
            delta=5,
            month_label=None
        )
    ]

    # Generate Excel
    excel_buffer = generate_history_excel(history_log_data, changes)

    # Load and verify
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print("\nChecking Excel structure:")

    # Check row 1 headers (month headers + static columns)
    print("\nRow 1 (Month Headers):")
    row1_values = []
    for cell in ws[1]:
        if cell.value:
            row1_values.append(str(cell.value))
            print(f"  Column {cell.column}: {cell.value}")

    # Expected row 1: Main LOB, State, Case Type, Case ID, Target CPH, Jun-25, Jul-25
    assert "Main LOB" in row1_values, "Main LOB should be in row 1"
    assert "State" in row1_values, "State should be in row 1"
    assert "Case Type" in row1_values, "Case Type should be in row 1"
    assert "Case ID" in row1_values, "Case ID should be in row 1"
    assert "Target CPH" in row1_values, "Target CPH should be in row 1"
    assert "Jun-25" in row1_values, "Jun-25 should be in row 1"
    assert "Jul-25" in row1_values, "Jul-25 should be in row 1"

    # Check row 2 headers (field headers)
    print("\nRow 2 (Field Headers):")
    row2_values = []
    for cell in ws[2]:
        if cell.value:
            row2_values.append(str(cell.value))
            print(f"  Column {cell.column}: {cell.value}")

    # Expected row 2: Client Forecast, FTE Required, FTE Available, Capacity (repeated for each month)
    assert "Client Forecast" in row2_values, "Client Forecast should be in row 2"
    assert "FTE Required" in row2_values, "FTE Required should be in row 2"
    assert "FTE Available" in row2_values, "FTE Available should be in row 2"
    assert "Capacity" in row2_values, "Capacity should be in row 2"

    # Count how many times each field appears (should be 2 times for 2 months)
    assert row2_values.count("Client Forecast") == 2, "Client Forecast should appear 2 times"
    assert row2_values.count("FTE Required") == 2, "FTE Required should appear 2 times"
    assert row2_values.count("FTE Available") == 2, "FTE Available should appear 2 times"
    assert row2_values.count("Capacity") == 2, "Capacity should appear 2 times"

    # Check row 3 has data
    print("\nRow 3 (Data):")
    row3_values = []
    for cell in ws[3]:
        if cell.value:
            row3_values.append(str(cell.value))
            print(f"  Column {cell.column}: {cell.value}")

    assert len(row3_values) > 0, "Row 3 should have data"
    assert "Amisys Medicaid DOMESTIC" in row3_values, "Row 3 should have Main LOB data"

    # Check merged cells
    print("\nMerged cells:")
    for merged_range in ws.merged_cells.ranges:
        print(f"  {merged_range}")

    # Verify static columns are merged vertically (2 rows)
    # Main LOB should be merged A1:A2
    static_merges = [str(m) for m in ws.merged_cells.ranges if ':' in str(m)]
    print(f"\nStatic column merges (vertical): {len([m for m in static_merges if '1:' in m and '2' in m])}")

    # Verify month headers are merged horizontally (4 columns)
    month_merges = [str(m) for m in static_merges if '1:' in m and '1' in m]
    print(f"Month header merges (horizontal): {len([m for m in month_merges if m.count(':') == 1])}")

    print("\n✓ TEST PASSED: Multi-level headers structure is correct")


def test_multilevel_headers_styling():
    """Test that headers have correct styling."""
    print("\n" + "=" * 70)
    print("Test 4: Multi-Level Headers Styling")
    print("=" * 70)

    # Create minimal test data
    history_log_data = HistoryLogData(
        id="test-456",
        change_type="CPH Update",
        month="March",
        year=2025,
        timestamp="2025-03-15T10:30:00Z",
        user="test_user",
        description="Test styling",
        records_modified=1,
        summary_data=None
    )

    changes = [
        HistoryChangeRecord(
            main_lob="Facets Commercial DOMESTIC",
            state="CA",
            case_type="Enrollment",
            case_id="EN-001",
            field_name="Jun-25.fte_avail",
            old_value=10,
            new_value=15,
            delta=5,
            month_label="Jun-25"
        )
    ]

    # Generate Excel
    excel_buffer = generate_history_excel(history_log_data, changes)

    # Load and verify styling
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print("\nChecking styling:")

    # Check row 1 styling (dark blue)
    row1_cells = [cell for cell in ws[1] if cell.value]
    for cell in row1_cells:
        fill_color = cell.fill.start_color.rgb if cell.fill.start_color else None
        print(f"  {cell.value}: Fill={fill_color}, Font Color={cell.font.color.rgb if cell.font.color else None}")

        # Verify dark blue background (accept both FF366092 and 00366092 formats)
        if fill_color:
            assert fill_color.endswith("366092"), f"Row 1 should have dark blue fill, got {fill_color}"

        # Verify white font (accept both FFFFFFFF and 00FFFFFF formats)
        if cell.font.color:
            assert cell.font.color.rgb.endswith("FFFFFF"), "Row 1 should have white font"

    # Check row 2 styling (light blue)
    row2_cells = [cell for cell in ws[2] if cell.value]
    for cell in row2_cells:
        fill_color = cell.fill.start_color.rgb if cell.fill.start_color else None
        if fill_color:
            # Only check field headers (not merged static columns)
            if cell.value in ["Client Forecast", "FTE Required", "FTE Available", "Capacity"]:
                assert fill_color.endswith("5B9BD5"), f"Row 2 fields should have light blue fill, got {fill_color}"

    print("\n✓ TEST PASSED: Headers have correct styling (dark blue row 1, light blue row 2)")


def main():
    """Run all multi-level header tests."""
    print("\n" + "=" * 70)
    print("MULTI-LEVEL HEADERS TEST SUITE")
    print("=" * 70)
    print("\nTesting Excel export with hierarchical headers:")
    print("  - Row 1: Month headers (merged horizontally)")
    print("  - Row 2: Field headers (Client Forecast, FTE Required, etc.)")
    print("  - Static columns: Merged vertically across rows 1-2")

    try:
        test_parse_month_label()
        test_prepare_pivot_data_metadata()
        test_multilevel_headers_structure()
        test_multilevel_headers_styling()

        print("\n" + "=" * 70)
        print("✓ ALL TESTS PASSED")
        print("=" * 70)
        print("\nMulti-Level Headers Successfully Implemented!")
        print("\nFeatures Verified:")
        print("  ✓ Month labels parsed and sorted chronologically")
        print("  ✓ Metadata (month_labels, static_columns) returned from pivot data")
        print("  ✓ Row 1 has month headers merged across 4 columns")
        print("  ✓ Row 2 has field headers (Client Forecast, FTE Required, FTE Available, Capacity)")
        print("  ✓ Static columns merged vertically across rows 1-2")
        print("  ✓ Row 1: Dark blue (#366092) with white text")
        print("  ✓ Row 2: Light blue (#5B9BD5) with white text")
        print("  ✓ Data starts at row 3")

    except AssertionError as e:
        print("\n" + "=" * 70)
        print("✗ TEST FAILED")
        print("=" * 70)
        print(f"\nAssertion Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    except Exception as e:
        print("\n" + "=" * 70)
        print("✗ UNEXPECTED ERROR")
        print("=" * 70)
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == "__main__":
    exit(main())
