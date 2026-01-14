"""
Diagnose Excel auto-repair issue.
This test will check for common issues that cause Excel to require repairs.
"""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    HistoryLogData,
    HistoryChangeRecord,
    generate_history_excel
)
from openpyxl import load_workbook
import tempfile


def diagnose_excel_structure():
    """Create Excel and check for issues that cause auto-repair."""
    print("\n" + "=" * 70)
    print("EXCEL AUTO-REPAIR DIAGNOSIS")
    print("=" * 70)

    # Create realistic test data
    history_log = HistoryLogData(
        id="ba-12345",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T14:30:00Z",
        user="system",
        description="Automated bench allocation for March 2025",
        records_modified=1,
        summary_data={
            'report_month': 'March',
            'report_year': 2025,
            'months': ['Jun-25'],
            'totals': {
                'Jun-25': {
                    'total_fte_available': {'old': 100, 'new': 105}
                }
            }
        }
    )

    # Create changes for all 4 fields
    changes = [
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.forecast",
            old_value=1000,
            new_value=1000,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.fte_req",
            old_value=20,
            new_value=20,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,
            delta=5,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1125,
            delta=125,
            month_label="Jun-25"
        )
    ]

    print("\n1. Generating Excel...")
    excel_buffer = generate_history_excel(history_log, changes)

    # Save to temp file
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        tmp.write(excel_buffer.getvalue())
        tmp_path = tmp.name

    print(f"   ✓ Saved to: {tmp_path}")

    # Load with openpyxl
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print("\n2. Checking Changes Sheet Structure:")
    print(f"   Total rows: {ws.max_row}")
    print(f"   Total columns: {ws.max_column}")

    # Check for None values in header rows
    print("\n3. Checking for None values in header rows:")
    issues = []

    for row_idx in [1, 2]:
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                issues.append(f"   ✗ Row {row_idx}, Column {col_idx} is None")

    if issues:
        print("   Issues found:")
        for issue in issues:
            print(issue)
    else:
        print("   ✓ No None values in header rows")

    # Check merged cell ranges
    print("\n4. Checking merged cell ranges:")
    merged_ranges = list(ws.merged_cells.ranges)
    print(f"   Total merged ranges: {len(merged_ranges)}")

    for i, mr in enumerate(merged_ranges, 1):
        print(f"   {i}. {mr}")

        # Check if merge range is valid (start <= end)
        if mr.min_col > mr.max_col or mr.min_row > mr.max_row:
            print(f"      ✗ INVALID: min > max")
            issues.append(f"Invalid merge range: {mr}")

        # Check if cells in range have values
        anchor_cell = ws.cell(row=mr.min_row, column=mr.min_col)
        if anchor_cell.value is None:
            print(f"      ✗ Anchor cell is None")
            issues.append(f"Anchor cell at {mr.min_row},{mr.min_col} is None")
        else:
            print(f"      ✓ Anchor value: {anchor_cell.value}")

    # Check data rows
    print("\n5. Checking data rows (row 3+):")
    for row_idx in range(3, min(ws.max_row + 1, 6)):  # Check first few data rows
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_values.append(str(cell.value) if cell.value is not None else "(empty)")
        print(f"   Row {row_idx}: {row_values}")

    # Check Summary sheet
    print("\n6. Checking Summary Sheet:")
    if 'Summary' in wb.sheetnames:
        ws_summary = wb['Summary']
        print(f"   ✓ Summary sheet exists")
        print(f"   Rows: {ws_summary.max_row}")
        print(f"   Columns: {ws_summary.max_column}")

        # Check for None values
        none_count = 0
        for row in ws_summary.iter_rows():
            for cell in row:
                if cell.value is None:
                    none_count += 1

        if none_count > 0:
            print(f"   Warning: {none_count} cells with None values")
    else:
        print("   ✗ Summary sheet missing")
        issues.append("Summary sheet missing")

    # Summary
    print("\n" + "=" * 70)
    if issues:
        print("✗ ISSUES FOUND THAT MAY CAUSE AUTO-REPAIR:")
        print("=" * 70)
        for issue in issues:
            print(f"  - {issue}")
    else:
        print("✓ NO OBVIOUS ISSUES DETECTED")
        print("=" * 70)
        print("\nThe file structure looks correct. If Excel still repairs it,")
        print("the issue may be related to:")
        print("  1. Excel version compatibility")
        print("  2. Specific cell formatting")
        print("  3. openpyxl version compatibility")

    print(f"\n\nFile saved at: {tmp_path}")
    print("Please open this file in Excel and report any specific error messages.")

    return tmp_path


if __name__ == "__main__":
    try:
        path = diagnose_excel_structure()
        print("\n✓ Diagnosis complete")
    except Exception as e:
        print(f"\n✗ Error during diagnosis: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
