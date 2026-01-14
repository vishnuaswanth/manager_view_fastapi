"""
Test saving Excel BEFORE merging to see if values are preserved.
"""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import zipfile
import xml.etree.ElementTree as ET


def test_before_and_after_merge():
    """Compare XML before and after merging."""
    print("\n" + "=" * 70)
    print("TESTING: VALUES BEFORE AND AFTER MERGE")
    print("=" * 70)

    # Create workbook
    wb = Workbook()
    ws = wb.active

    # Write same value to cells that will be merged
    print("\n1. Writing 'Main LOB' to cells A1 and A2...")
    ws.cell(row=1, column=1, value="Main LOB")
    ws.cell(row=2, column=1, value="Main LOB")

    # Save BEFORE merge
    print("\n2. Saving BEFORE merge...")
    buffer_before = BytesIO()
    wb.save(buffer_before)
    buffer_before.seek(0)

    # Check XML before merge
    with zipfile.ZipFile(buffer_before, 'r') as zip_ref:
        sheet_xml = zip_ref.read('xl/worksheets/sheet1.xml')
        root = ET.fromstring(sheet_xml)
        ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

        print("   XML content (cells A1, A2):")
        for row_elem in root.findall('.//{%s}row' % ns[''], ns):
            for cell_elem in row_elem.findall('.//{%s}c' % ns[''], ns):
                cell_ref = cell_elem.get('r')
                if cell_ref in ['A1', 'A2']:
                    # Check for inline string
                    is_elem = cell_elem.find('.//{%s}is' % ns[''], ns)
                    if is_elem is not None:
                        t_elem = is_elem.find('.//{%s}t' % ns[''], ns)
                        if t_elem is not None:
                            print(f"     {cell_ref}: <is><t>{t_elem.text}</t></is> ✓")
                    else:
                        print(f"     {cell_ref}: NO inline string value ✗")

    # Now merge
    print("\n3. Merging A1:A2...")
    ws.merge_cells('A1:A2')

    # Save AFTER merge
    print("\n4. Saving AFTER merge...")
    buffer_after = BytesIO()
    wb.save(buffer_after)
    buffer_after.seek(0)

    # Check XML after merge
    with zipfile.ZipFile(buffer_after, 'r') as zip_ref:
        sheet_xml = zip_ref.read('xl/worksheets/sheet1.xml')
        root = ET.fromstring(sheet_xml)
        ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

        print("   XML content (cells A1, A2):")
        for row_elem in root.findall('.//{%s}row' % ns[''], ns):
            for cell_elem in row_elem.findall('.//{%s}c' % ns[''], ns):
                cell_ref = cell_elem.get('r')
                if cell_ref in ['A1', 'A2']:
                    # Check for inline string
                    is_elem = cell_elem.find('.//{%s}is' % ns[''], ns)
                    if is_elem is not None:
                        t_elem = is_elem.find('.//{%s}t' % ns[''], ns)
                        if t_elem is not None:
                            print(f"     {cell_ref}: <is><t>{t_elem.text}</t></is> ✓")
                    else:
                        cell_type = cell_elem.get('t')
                        print(f"     {cell_ref}: NO inline string value (type={cell_type}) ✗")

        # Check merge cells definition
        merged_cells = root.find('.//{%s}mergeCells' % ns[''], ns)
        if merged_cells:
            print("\n   Merged cells:")
            for mc in merged_cells.findall('.//{%s}mergeCell' % ns[''], ns):
                print(f"     - {mc.get('ref')}")

    print("\n" + "=" * 70)
    print("CONCLUSION:")
    print("=" * 70)
    print("If values disappear AFTER merge, openpyxl is removing them.")
    print("This would be an openpyxl bug/limitation.")


if __name__ == "__main__":
    try:
        test_before_and_after_merge()
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
