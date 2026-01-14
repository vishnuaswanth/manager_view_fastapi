"""
Test the full API flow for Excel generation to identify corruption source.
"""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    HistoryLogData,
    HistoryChangeRecord,
    generate_history_excel
)
from io import BytesIO
import tempfile


def test_full_api_flow():
    """Simulate the full API flow: generate Excel, return BytesIO, save to file."""
    print("\n" + "=" * 70)
    print("TESTING FULL API FLOW")
    print("=" * 70)

    # Create realistic bench allocation data
    history_log = HistoryLogData(
        id="ba-test-001",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T14:30:00Z",
        user="system",
        description="Test bench allocation",
        records_modified=2,
        summary_data={
            'report_month': 'March',
            'report_year': 2025,
            'months': ['Jun-25', 'Jul-25'],
            'totals': {
                'Jun-25': {
                    'total_fte_available': {'old': 100, 'new': 105},
                    'total_capacity': {'old': 5000, 'new': 5250}
                },
                'Jul-25': {
                    'total_fte_available': {'old': 100, 'new': 110},
                    'total_forecast': {'old': 6000, 'new': 6000}
                }
            }
        }
    )

    # Create multiple records with multiple months
    changes = []

    # Record 1: Jun-25
    for field in ['forecast', 'fte_req', 'fte_avail', 'capacity']:
        changes.append(HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name=f"Jun-25.{field}",
            old_value=1000 if field == 'forecast' else 20,
            new_value=1000 if field == 'forecast' else (25 if field == 'fte_avail' else (1125 if field == 'capacity' else 20)),
            delta=0 if field in ['forecast', 'fte_req'] else (5 if field == 'fte_avail' else 125),
            month_label="Jun-25"
        ))

    # Record 2: Jul-25
    for field in ['forecast', 'fte_req', 'fte_avail', 'capacity']:
        changes.append(HistoryChangeRecord(
            main_lob="Facets Commercial DOMESTIC",
            state="CA",
            case_type="Enrollment",
            case_id="EN-002",
            field_name=f"Jul-25.{field}",
            old_value=2000 if field == 'forecast' else 40,
            new_value=2000 if field == 'forecast' else (50 if field == 'fte_avail' else (2500 if field == 'capacity' else 40)),
            delta=0 if field in ['forecast', 'fte_req'] else (10 if field == 'fte_avail' else 500),
            month_label="Jul-25"
        ))

    print("\n1. Calling generate_history_excel()...")
    excel_buffer = generate_history_excel(history_log, changes)

    print(f"   ✓ Buffer size: {excel_buffer.tell()} bytes")
    print(f"   ✓ Buffer type: {type(excel_buffer)}")

    # Simulate what FastAPI StreamingResponse does
    print("\n2. Simulating API response (seek to 0, read all)...")
    excel_buffer.seek(0)
    excel_content = excel_buffer.read()
    print(f"   ✓ Content size: {len(excel_content)} bytes")

    # Save to file (as user would download)
    print("\n3. Saving to file (simulating download)...")
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        tmp.write(excel_content)
        tmp_path = tmp.name

    print(f"   ✓ Saved to: {tmp_path}")

    # Try to open with openpyxl to verify structure
    print("\n4. Verifying file structure with openpyxl...")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(tmp_path)
        print("   ✓ openpyxl can open file")

        ws_changes = wb['Changes']
        print(f"   ✓ Changes sheet: {ws_changes.max_row} rows, {ws_changes.max_column} cols")

        ws_summary = wb['Summary']
        print(f"   ✓ Summary sheet: {ws_summary.max_row} rows, {ws_summary.max_column} cols")

        # Check merged cells
        merged_ranges = list(ws_changes.merged_cells.ranges)
        print(f"   ✓ Merged cell ranges: {len(merged_ranges)}")

        # Check for None values in headers
        none_count_r1 = sum(1 for col in range(1, ws_changes.max_column + 1)
                           if ws_changes.cell(row=1, column=col).value is None)
        none_count_r2 = sum(1 for col in range(1, ws_changes.max_column + 1)
                           if ws_changes.cell(row=2, column=col).value is None)

        print(f"   Row 1 None values: {none_count_r1}")
        print(f"   Row 2 None values: {none_count_r2}")

        if none_count_r1 > 0 or none_count_r2 > 0:
            print("   Note: None values in merged cells are normal")

    except Exception as e:
        print(f"   ✗ Failed to open: {e}")
        import traceback
        traceback.print_exc()
        return None

    print("\n" + "=" * 70)
    print("✓ FULL API FLOW TEST COMPLETE")
    print("=" * 70)
    print(f"\nFile location: {tmp_path}")
    print("\nPlease open this file in Excel and check:")
    print("  1. Does Excel show a repair message?")
    print("  2. If yes, what is the EXACT error message?")
    print("  3. Does the file open correctly after repair (if needed)?")
    print("  4. Are all headers and data correct?")

    return tmp_path


if __name__ == "__main__":
    try:
        path = test_full_api_flow()
        if path:
            print(f"\n✓ Test completed: {path}")
    except Exception as e:
        print(f"\n✗ Test failed: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
