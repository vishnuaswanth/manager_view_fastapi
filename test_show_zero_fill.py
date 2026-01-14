"""
Demonstrate the 0-fill behavior for missing fields.
"""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    HistoryLogData,
    HistoryChangeRecord,
    generate_history_excel
)
from openpyxl import load_workbook
import tempfile


def test_zero_fill_demonstration():
    """Create an Excel file showing 0-fill for missing fields."""
    print("\n" + "=" * 70)
    print("DEMONSTRATION: Missing Fields Filled with 0")
    print("=" * 70)

    history_log = HistoryLogData(
        id="demo-001",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T14:30:00Z",
        user="system",
        description="Demo: Partial field changes",
        records_modified=2,
        summary_data=None
    )

    # Record 1: Only FTE Available and Capacity changed (Forecast and FTE Required missing)
    # Record 2: All fields present
    changes = [
        # Record 1: TX - Only 2 fields
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,
            delta=5,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1125,
            delta=125,
            month_label="Jun-25"
        ),

        # Record 2: CA - All 4 fields
        HistoryChangeRecord(
            main_lob="Facets Commercial DOMESTIC",
            state="CA",
            case_type="Enrollment",
            case_id="EN-002",
            field_name="Jun-25.forecast",
            old_value=2000,
            new_value=2000,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Facets Commercial DOMESTIC",
            state="CA",
            case_type="Enrollment",
            case_id="EN-002",
            field_name="Jun-25.fte_req",
            old_value=40,
            new_value=40,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Facets Commercial DOMESTIC",
            state="CA",
            case_type="Enrollment",
            case_id="EN-002",
            field_name="Jun-25.fte_avail",
            old_value=40,
            new_value=50,
            delta=10,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Facets Commercial DOMESTIC",
            state="CA",
            case_type="Enrollment",
            case_id="EN-002",
            field_name="Jun-25.capacity",
            old_value=2000,
            new_value=2500,
            delta=500,
            month_label="Jun-25"
        ),
    ]

    print("\n1. Creating Excel with partial fields...")
    excel_buffer = generate_history_excel(history_log, changes)

    # Save to file
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        tmp.write(excel_buffer.getvalue())
        tmp_path = tmp.name

    print(f"   ✓ Saved to: {tmp_path}")

    # Load and display
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print("\n2. Excel Content:")
    print("\n   Headers:")
    print("   Row 1:", [ws.cell(row=1, column=i).value for i in range(1, 9)])
    print("   Row 2:", [ws.cell(row=2, column=i).value for i in range(1, 9)])

    print("\n   Data Rows:")
    for row_num in range(3, ws.max_row + 1):
        row_values = [ws.cell(row=row_num, column=i).value for i in range(1, 9)]
        print(f"   Row {row_num}: {row_values}")

    print("\n3. Analysis:")

    # Check Record 1 (TX - partial fields)
    print("\n   Record 1 (TX - Only FTE Avail & Capacity changed):")
    print(f"     Client Forecast: {ws.cell(row=3, column=5).value} ← Filled with 0 (was missing)")
    print(f"     FTE Required:    {ws.cell(row=3, column=6).value} ← Filled with 0 (was missing)")
    print(f"     FTE Available:   {ws.cell(row=3, column=7).value} ← Has data")
    print(f"     Capacity:        {ws.cell(row=3, column=8).value} ← Has data")

    # Check Record 2 (CA - all fields)
    print("\n   Record 2 (CA - All fields present):")
    print(f"     Client Forecast: {ws.cell(row=4, column=5).value} ← Has data")
    print(f"     FTE Required:    {ws.cell(row=4, column=6).value} ← Has data")
    print(f"     FTE Available:   {ws.cell(row=4, column=7).value} ← Has data")
    print(f"     Capacity:        {ws.cell(row=4, column=8).value} ← Has data")

    print("\n" + "=" * 70)
    print("✓ DEMONSTRATION COMPLETE")
    print("=" * 70)
    print(f"\nOpen the file to see the 0-fill behavior:")
    print(f"  open {tmp_path}")
    print("\nMissing fields (Forecast, FTE Required) for TX record show as 0")

    return tmp_path


if __name__ == "__main__":
    try:
        path = test_zero_fill_demonstration()
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
