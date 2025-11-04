# Standard library imports
import sys
import os
import re
import copy
import traceback
import logging
from datetime import datetime
from typing import List, Dict, Tuple

# Third-party imports
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Local application imports - settings
from code.settings import BASE_DIR, MODE, SQLITE_DATABASE_URL, MSSQL_DATABASE_URL

# Local application imports - core utilities
from code.logics.core_utils import PreProcessing, CoreUtils

# Local application imports - database models
from code.logics.db import ProdTeamRosterModel, AllocationReportsModel

# Local application imports - utility modules
from code.logics.month_config_utils import get_specific_config
from code.logics.allocation_tracker import start_execution, update_status, complete_execution
from code.logics.export_utils import (
    get_latest_or_requested_dataframe,
    update_forecast_data,
    get_forecast_months_list,
    get_all_model_dataframes_dict,
    get_calculations_data
)
from code.logics.summary_utils import update_summary_data
from code.logics.manager_view import parse_main_lob

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

pd.set_option('future.no_silent_downcasting', True)

# Determine database URL based on mode
if MODE.upper() == "DEBUG":
    DATABASE_URL = SQLITE_DATABASE_URL
elif MODE.upper() == "PRODUCTION":
    DATABASE_URL = MSSQL_DATABASE_URL
else:
    raise ValueError("Invalid MODE specified in config.")

# Initialize CoreUtils instance (singleton for this module)
core_utils = CoreUtils(DATABASE_URL)

# Set current path
curpth = os.path.join(BASE_DIR, 'logics')

# Input and output file paths
input_files = {
    'skilling': os.path.join(curpth, "data","Input", "Centene Modified Roster-2.xlsm"),
    'target': os.path.join(curpth, "data",'constants','calculations.xlsx'),
    'vendor_template': os.path.join(curpth, "data","Input", "NTT_Capacity Roster Template.xlsx"),
    # No longer needed with ResourceAllocator:
    # 'combinations': os.path.join(curpth, "data",'constants',"combinations_2.xlsx")
}
output_file = os.path.join(curpth, "result.xlsx")



# Load variables and months
req_vars_df = pd.read_excel(input_files['target'], sheet_name="Sheet1")
# req_months_df = pd.read_excel(input_files['target'], sheet_name="Sheet2")
occupancy = req_vars_df['Occupancy'].iloc[0]
shrinkage = req_vars_df['Shrinkage'].iloc[0]
workhours = req_vars_df['Work hours'].iloc[0]
# month_headers = req_months_df['Months'].tolist()
month_with_days = dict(zip(req_vars_df['months'], req_vars_df['No.of days occupancy']))

# Legacy code - no longer needed with ResourceAllocator
# combination_list = [eval(x) for x in combinations_df['combination'] if 'nan' not in x]
# state_with_worktype_volume_dict = {}

# Helper functions
def get_value(row, month, filetype, df:pd.DataFrame=None):
    if df is None or getattr(df, 'empty', True):
        return 0

    file_name = row[('Centene Capacity plan', 'Main LOB')]
    state = row[('Centene Capacity plan', 'State')]
    work_type = row[('Centene Capacity plan', 'Case type')]
    # file_paths = {
    #     'nonmmp': os.path.join(curpth, "data", 'constants',"medicare_medicaid_nonmmp", f"{file_name}.xlsx"),
    #     'mmp': os.path.join(curpth, "data", 'constants',"medicare_medicaid_mmp", f"{file_name}.xlsx"),
    #     'summary': os.path.join(curpth, "data", 'constants',"medicare_medicaid_summary", f"{file_name}-summary.xlsx")
    # }
    try:
        if filetype == 'medicare_medicaid_nonmmp':
            # df = pd.read_excel(file_paths['nonmmp'], header=[0, 1, 2])
            # df.columns = df.columns.map(lambda x: tuple(i if 'Unnamed' not in str(i) else '' for i in x))
            filtered_df = df[(df[('', '', 'State')] == state) & (df[('', '', 'Month')] == month)]
            if not filtered_df.empty:
                for col in filtered_df.columns:
                    if "WFM TO PROVIDE" in col[0].upper() and work_type in col:
                        return filtered_df[col].values[0]
            return 0
        elif filetype == 'medicare_medicaid_mmp':
            # df = pd.read_excel(file_paths['mmp'], header=[0, 1])
            filtered_df = df[(df[('State', 'State')] == state) & (df[('Month', 'Month')] == month)]
            if not filtered_df.empty:
                for col in filtered_df.columns:
                    if work_type == col[0]:
                        return filtered_df[col].values[0]
            return 0
        elif filetype == 'medicare_medicaid_summary':
            # df = pd.read_excel(file_paths['summary'], header=[0, 1, 2, 3])
            # df.columns = df.columns.map(lambda x: tuple(datetime.strptime(i, "%Y-%m-%d %H:%M:%S").strftime('%B') if ":" in i and "unnamed" not in i.lower() else i for i in x))
            filtered_df = df[df[(file_name, "Vendor Eligible Forecast (WFM)", "Work Type", "")] == work_type]
            if not filtered_df.empty:
                return filtered_df[(file_name, "Vendor Eligible Forecast (WFM)", month, "")].values[0]
            return 0
    except KeyError as e:
        logging.warning(f"Missing month column '{month}' in get_value for {file_name}, {filetype}, returning 0")
        return 0
    except Exception as e:
        logging.warning(f"Error in get_value for {file_name}, {filetype}, {month}: {e}")
        return 0

class Calculations():
    def __init__(self, data_month: str = None, data_year: int = None) -> None:
        self.target_cph: pd.DataFrame = pd.DataFrame()
        self.month_data: pd.DataFrame = pd.DataFrame()

        # Store month and year for database lookups
        self.data_month = data_month
        self.data_year = data_year

        # Cache for month configurations to avoid repeated DB queries
        self._config_cache: Dict[Tuple[str, int, str], Dict] = {}

        calculations = get_calculations_data()
        self.month_data = calculations.get("month_data", pd.DataFrame)
        self.target_cph = calculations.get("target_cph", pd.DataFrame)

    def get_config_for_worktype(self, month: str, year: int, work_type: str) -> Dict:
        """
        Get configuration parameters for a specific month, year, and work type.

        Uses database-backed MonthConfigurationModel for work-type-specific parameters.
        CRITICAL: Raises ValueError if configuration not found - allocation cannot proceed
        without accurate month-specific parameters.

        Args:
            month: Month name (e.g., "January")
            year: Year (e.g., 2025)
            work_type: "Domestic" or "Global"

        Returns:
            Dictionary with keys: working_days, occupancy, shrinkage, work_hours

        Raises:
            ValueError: If month configuration not found in database
        """
        # Check cache first
        cache_key = (month, year, work_type)
        if cache_key in self._config_cache:
            return self._config_cache[cache_key]

        # Try to get from database
        config = get_specific_config(month, year, work_type)

        if config:
            result = {
                'working_days': config['working_days'],
                'occupancy': config['occupancy'],
                'shrinkage': config['shrinkage'],
                'work_hours': config['work_hours']
            }
            # Cache the result
            self._config_cache[cache_key] = result
            logger.debug(f"Loaded config from DB for {month} {year} ({work_type}): {result}")
            return result
        else:
            # CRITICAL: Month configuration missing - cannot proceed
            error_msg = (
                f"CRITICAL: No month configuration found for {month} {year} ({work_type}). "
                f"Allocation cannot proceed with accurate parameters. "
                f"Please add the missing configuration via POST /api/month-config endpoint. "
                f"Both Domestic and Global configurations are required for each month-year."
            )
            logger.error(error_msg)
            raise ValueError(error_msg)


    def get_target_cph(self, row):
        lob = row[('Centene Capacity plan', 'Main LOB')]
        worktype = row[('Centene Capacity plan', 'Case type')]
        logger.debug(f"~~ ENTER get_target_cph: lob={lob!r}, worktype={worktype!r}")
        target_df = self.target_cph
        # filtered_target_df = self.target_cph[(self.target_cph['Case type'].str.lower() == worktype.lower()) & (self.target_cph['Main LOB'].str.strip() == lob.strip())]
        filtered_target_df = target_df[
            (target_df['Main LOB'].str.strip().str.lower() == lob.strip().lower()) &
            (target_df['Case type'].str.strip().str.lower() == worktype.strip().lower())
        ]
        # logger.debug(f"filtered target df head - {filtered_target_df.head()}")
        return filtered_target_df['Target CPH'].iloc[0] if not filtered_target_df.empty else 0



def get_fte_required(row, month, calculations: Calculations):
    """
    Calculate FTE Required using work-type-specific configuration parameters.

    Determines work type (Domestic/Global) from LOB parsing, with special handling
    for OIC Volumes where work type is extracted from the Case Type column.

    Formula: Volume / (Target_CPH * WorkHours * Occupancy * (1 - Shrinkage) * WorkingDays)
    """
    try:
        target_cph = row[('Centene Capacity plan', 'Target CPH')]
        month_value = row[('Client Forecast', month)]
        main_lob = row[('Centene Capacity plan', 'Main LOB')]
        case_type = row[('Centene Capacity plan', 'Case type')]

        # Determine work type based on LOB parsing
        parsed_lob = parse_main_lob(main_lob)
        lob_locality = parsed_lob.get('locality', '')

        # SPECIAL CASE: OIC Volumes - locality determined from worktype column
        is_oic_volumes = 'oic' in str(main_lob).lower() and 'volumes' in str(main_lob).lower()
        if is_oic_volumes:
            case_type_lower = str(case_type).lower()
            work_type = 'Domestic' if 'domestic' in case_type_lower else 'Global'
        else:
            # Normalize locality to Domestic/Global
            work_type = 'Domestic' if 'domestic' in str(lob_locality).lower() else 'Global'

        # Get work-type-specific configuration
        year = calculations.data_year if calculations.data_year else datetime.now().year  # Fallback to current year
        config = calculations.get_config_for_worktype(month, year, work_type)

        no_of_days = config['working_days']
        occupancy = config['occupancy']
        shrinkage = config['shrinkage']
        workhours = config['work_hours']

        if target_cph == 0 or no_of_days == 0:
            return 0

        fte_required = month_value / (target_cph * workhours * occupancy * (1 - shrinkage) * no_of_days)
        return fte_required

    except (KeyError, TypeError) as e:
        logging.warning(f"Missing month data for {month} in get_fte_required: {e}, returning 0")
        return 0
    except ValueError:
        # ValueError indicates missing month config - must propagate to halt execution
        raise
    except Exception as e:
        logging.error(f"Error in get_fte_required for {month}: {e}", exc_info=True)
        return 0

def get_temp_casetype(casetype):
    casetype = str(casetype)
    if not casetype or casetype == 'nan':
        return ''
    ct = casetype.split("-")[0].lower()
    return {'app': 'appeal', 'omn': 'omni'}.get(ct, ct)


def normalize_locality(locality_str: str) -> str:
    """
    Normalize locality to Domestic or Global (case-insensitive).

    Handles variations like 'Domestic', 'domestic', '(Domestic)', '(domestic)', etc.

    Args:
        locality_str: Locality string from vendor Location or parsed LOB

    Returns:
        'Domestic' if contains 'domestic' (case-insensitive), else 'Global'

    Examples:
        >>> normalize_locality('Domestic')
        'Domestic'
        >>> normalize_locality('(domestic)')
        'Domestic'
        >>> normalize_locality('Global')
        'Global'
        >>> normalize_locality('')
        'Global'
    """
    if not locality_str:
        return 'Global'

    locality_lower = str(locality_str).lower().strip()
    # Remove parentheses for matching
    locality_clean = locality_lower.replace('(', '').replace(')', '')

    if 'domestic' in locality_clean:
        return 'Domestic'
    return 'Global'


class ResourceAllocator:
    """
    Fast resource allocation system using vocabulary-driven exact matching.

    Features:
    - Pre-computes all resource buckets upfront (O(n) initialization)
    - Uses vocabulary from demand to parse vendor skills
    - Greedy longest-match-first algorithm prevents substring issues
    - Handles multi-skilled vendors with priority: single-skill → multi-skill
    - Tracks allocation history for comprehensive reporting
    """

    def __init__(self, vendor_df: pd.DataFrame, output_df: pd.DataFrame, month_headers: List[str]):
        """
        Initialize allocator with pre-computed buckets.

        Args:
            vendor_df: DataFrame with vendor resources (must have PrimaryPlatform, State, NewWorkType)
            output_df: DataFrame with demand (must have worktype column)
            month_headers: List of month names for allocation
        """
        self.month_headers = month_headers
        self.allocation_history = []

        # Store reference to original vendor_df for report generation
        self.vendor_df_original = vendor_df.copy()

        # Reverse lookup index: vendor_df_idx -> {month: allocation_details}
        # Enables O(1) lookup during report generation
        self.vendor_allocations = {}

        # Extract valid states from demand
        self.valid_states = self._extract_valid_states(output_df)
        logger.info(f"Valid states from demand: {sorted(self.valid_states)}")

        # Build vocabulary from demand (sorted longest-first)
        self.worktype_vocab = self._build_vocabulary(output_df)
        logger.info(f"Built vocabulary with {len(self.worktype_vocab)} unique worktypes")
        logger.info(f"Sample worktypes: {self.worktype_vocab[:5]}")

        # Pre-compile regex for performance
        self.whitespace_pattern = re.compile(r'\s+')

        # Clean and expand vendor data by state
        vendor_df_clean = self._clean_and_expand_vendor_states(vendor_df)
        logger.info(f"Cleaned vendor data: {vendor_df_clean.shape[0]} records after state expansion")

        # Parse vendors and build buckets
        self.buckets = self._initialize_buckets(vendor_df_clean)

        # Store initial state for reporting
        self.initial_state = self._snapshot_state()
        total_vendor_instances = sum(len(vendors) for vendors in self.buckets.values())
        logger.info(f"Initialized allocator with {total_vendor_instances} total vendor-month instances across {len(self.buckets)} (platform, month, skillset) combinations")

        # Debug: Show sample bucket structure
        if self.buckets:
            sample_key = list(self.buckets.keys())[0]
            sample_vendors = self.buckets[sample_key]
            logger.info(f"Sample bucket key: {sample_key}")
            logger.info(f"Sample bucket vendor count: {len(sample_vendors)}")
            if sample_vendors:
                logger.info(f"Sample vendor states: {sample_vendors[0]['states']}")

        # Export buckets to Excel for debugging
        self._export_buckets_to_excel()

    def _extract_valid_states(self, output_df: pd.DataFrame) -> set:
        """
        Extract valid state codes from demand DataFrame.

        Returns:
            set: Valid state codes (including "N/A")
        """
        states = output_df[('Centene Capacity plan', 'State')].unique()
        valid_states = {
            str(state).strip().upper()
            for state in states
            if state and str(state).lower() not in {'nan', 'none', ''}
        }
        return valid_states

    def _clean_and_expand_vendor_states(self, vendor_df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean vendor State column and parse as list of states.

        CRITICAL: A vendor with "FL GA AR" is ONE resource that can work in FL, GA, or AR.
        They should only be counted ONCE, not multiple times!

        State mapping logic:
        - Parse multi-state strings like "FL GA AR" → keep as list [FL, GA, AR]
        - Filter: Keep matched states, convert unmatched to N/A
        - Store as StateList column for allocation matching

        Example:
          Demand states: [FL, GA, MI, N/A]
          Vendor state: "FL GA AR"
          Result: ONE record with StateList = [FL, GA, N/A]
                  (FL matches, GA matches, AR → N/A)

        Returns:
            DataFrame: Vendor data with StateList column (list of states vendor can work in)
        """
        vendor_df = vendor_df.copy()

        # Common US state codes (2-letter) for validation
        us_state_pattern = re.compile(r'^[A-Z]{2}$')

        # Get specific states (excluding N/A)
        specific_demand_states = self.valid_states - {'N/A'}

        def parse_states(state_str):
            """
            Parse state string into list of valid states.

            IMPORTANT: Every vendor can be used for N/A demands, so we ALWAYS add 'N/A' to StateList.
            This makes it explicit that vendor is available when demand has state='N/A'.
            """
            state_str = str(state_str).strip().upper()

            if not state_str or state_str in {'NAN', 'NONE', ''}:
                return ['N/A']  # No state info → N/A only

            # Split by whitespace
            state_tokens = state_str.split()

            parsed_states = []
            for token in state_tokens:
                if us_state_pattern.match(token):
                    # Valid 2-letter code
                    if token in specific_demand_states:
                        parsed_states.append(token)  # Matched state
                    else:
                        # Unmatched valid state code → don't add to list
                        # It will be available via N/A anyway
                        pass
                else:
                    # Invalid code → ignore (will be available via N/A)
                    pass

            # Remove duplicates while preserving order
            seen = set()
            unique_states = []
            for s in parsed_states:
                if s not in seen:
                    seen.add(s)
                    unique_states.append(s)

            # ALWAYS add 'N/A' - every vendor can fulfill N/A demands
            if 'N/A' not in unique_states:
                unique_states.append('N/A')

            return unique_states

        vendor_df['StateList'] = vendor_df['State'].apply(parse_states)

        logger.info(f"Parsed states for {len(vendor_df)} vendor records")
        logger.info(f"Sample StateList: {vendor_df['StateList'].head().tolist()}")

        # Debug: Count multi-state vendors
        multi_state_count = vendor_df['StateList'].apply(lambda x: len(x) > 1).sum()
        logger.info(f"Vendors with multiple states: {multi_state_count}")

        return vendor_df

    def _build_vocabulary(self, output_df: pd.DataFrame) -> List[str]:
        """
        Extract unique worktypes from demand DataFrame, sorted by length (longest first).

        Critical: Longest strings MUST be checked first to avoid substring matching issues.
        Example: "FTC-Basic/Non MMP" must be checked before "FTC" or "FTC Basic"
        """
        worktypes = output_df[('Centene Capacity plan', 'Case type')].unique()

        # Clean and filter vocabulary
        vocab = {
            str(wt).strip().lower()
            for wt in worktypes
            if wt and str(wt).lower() not in {'nan', 'none', ''}
        }

        # Sort by length DESC (longest first), then alphabetically for deterministic behavior
        return sorted(vocab, key=lambda x: (-len(x), x))

    def _normalize_text(self, text: str) -> str:
        """
        Normalize whitespace: collapse multiple spaces/tabs to single space, strip.

        Example: "FTC  ADJ" → "FTC ADJ"
        """
        if not text or str(text).lower() == 'nan':
            return ''
        # Replace multiple whitespace characters with single space, then strip
        return self.whitespace_pattern.sub(' ', str(text).strip())

    def _parse_vendor_skills(self, newworktype_str: str) -> frozenset:
        """
        Parse vendor NewWorkType by matching against vocabulary using greedy longest-match-first.

        Algorithm:
        1. Normalize whitespace and lowercase
        2. Find longest vocabulary term in remaining text
        3. Add to matched_skills, remove from text, re-normalize
        4. Repeat until no matches found

        Duplicates are automatically handled via set - if the same skill appears multiple times,
        it will only be included once in the result.

        Examples:
            Input: "FTC-Basic/Non MMP  ADJ-COB NON MMP" (note double space)
            Vocab: ["ftc-basic/non mmp", "adj-cob non mmp", "ftc", "adj", ...]
            Output: frozenset({'ftc-basic/non mmp', 'adj-cob non mmp'})

            Input: "FTC ADJ FTC" (duplicate FTC)
            Output: frozenset({'ftc', 'adj'})  # Deduplicates automatically
        """
        if not newworktype_str:
            return frozenset()

        # Step 1: Normalize and lowercase
        text = self._normalize_text(newworktype_str).lower()

        # Step 2: Greedy matching
        matched_skills = set()  # Use set for automatic deduplication

        while text:
            matched_any = False

            # Check each vocab term (already sorted longest-first)
            for vocab_term in self.worktype_vocab:
                if vocab_term in text:
                    matched_skills.add(vocab_term)  # Add to set (deduplicates automatically)
                    # Remove matched term and re-normalize
                    text = text.replace(vocab_term, ' ', 1)
                    text = self._normalize_text(text)
                    matched_any = True
                    break  # Start over from beginning of vocab (longest-first)

            if not matched_any:
                # No more vocabulary matches, stop
                # (remaining text contains only unknown/non-demand skills)
                break

        return frozenset(matched_skills)

    def _initialize_buckets(self, vendor_df: pd.DataFrame) -> dict:
        """
        Pre-compute all resource buckets grouped by (platform, month, skillset).

        CRITICAL: Vendors are stored with their StateList to avoid double-counting.
        Each vendor can only be allocated ONCE, even if they can work in multiple states.

        Returns:
            dict: {(platform, month, skillset): [list of vendor records with StateList]}
        """
        # Parse vendor skills
        vendor_df = vendor_df.copy()

        # Normalize platform: extract first word and uppercase for case-insensitive matching
        # Example: "Amisys CROP" → "AMISYS", "amisys" → "AMISYS"
        vendor_df['PlatformNormalized'] = vendor_df['PrimaryPlatform'].apply(
            lambda x: str(x).strip().split()[0].upper() if x and str(x).lower() != 'nan' else ''
        )

        logger.info(f"Parsing skills for {len(vendor_df)} vendor records...")
        vendor_df['ParsedSkills'] = vendor_df['NewWorkType'].apply(self._parse_vendor_skills)

        # Debug: Show parsing results
        logger.info(f"Sample vendor PrimaryPlatform: {vendor_df['PrimaryPlatform'].head().tolist()}")
        logger.info(f"Sample vendor PlatformNormalized: {vendor_df['PlatformNormalized'].head().tolist()}")
        logger.info(f"Sample vendor StateList: {vendor_df['StateList'].head().tolist()}")
        logger.info(f"Sample vendor NewWorkType: {vendor_df['NewWorkType'].head().tolist()}")
        logger.info(f"Sample ParsedSkills: {vendor_df['ParsedSkills'].head().tolist()}")

        # Filter out vendors with no recognized skills
        before_filter = len(vendor_df)
        vendor_df = vendor_df[vendor_df['ParsedSkills'].apply(len) > 0]
        after_filter = len(vendor_df)
        logger.info(f"Filtered vendors: {before_filter} → {after_filter} (removed {before_filter - after_filter} with no recognized skills)")

        if vendor_df.empty:
            logger.error("No vendors with recognized skills! Check worktype vocabulary matching.")
            return {}

        # Debug: Show platforms
        logger.info(f"Unique platforms (normalized) in vendor data: {sorted(vendor_df['PlatformNormalized'].unique())}")

        # Normalize vendor Location field
        vendor_df['LocationNormalized'] = vendor_df['Location'].apply(normalize_locality)
        logger.info(f"Unique locations (normalized) in vendor data: {sorted(vendor_df['LocationNormalized'].unique())}")

        # Create buckets: (platform, location, month, skillset) → list of vendor IDs with StateList
        # We'll track vendors by their index to prevent double-counting
        buckets = {}
        vendor_df['VendorID'] = range(len(vendor_df))  # Unique ID for each vendor

        for month in self.month_headers:
            month_normalized = str(month).strip().title()

            for idx, row in vendor_df.iterrows():
                platform = row['PlatformNormalized']
                location = row['LocationNormalized']
                skillset = row['ParsedSkills']
                vendor_id = row['VendorID']
                state_list = row['StateList']

                key = (platform, location, month_normalized, skillset)

                if key not in buckets:
                    buckets[key] = []

                # Store vendor with their state list and original DataFrame index
                buckets[key].append({
                    'vendor_id': vendor_id,
                    'vendor_df_idx': idx,  # Original DataFrame index for report lookup
                    'states': state_list,
                    'allocated': False  # Track if this vendor has been allocated
                })

        logger.info(f"Created buckets for {len(buckets)} (platform, location, month, skillset) combinations")
        total_vendors = sum(len(v) for v in buckets.values())
        logger.info(f"Total vendor-month instances: {total_vendors}")

        return buckets

    def _snapshot_state(self) -> dict:
        """Create deep copy of current bucket state for reporting."""
        import copy
        return copy.deepcopy(self.buckets)

    def generate_buckets_summary(self) -> tuple[pd.DataFrame, pd.DataFrame]:
        """
        Generate bucket summary data (without Excel export).

        Returns:
            tuple: (summary_df, details_df)
                - summary_df: Overview of all buckets with counts
                - details_df: Full vendor details for each bucket
        """
        summary_data = []
        details_data = []

        for (platform, location, month, skillset), vendors in sorted(self.buckets.items()):
            # Convert skillset to readable string
            skills_str = ' + '.join(sorted(skillset))

            # Get all unique states from vendors in this bucket
            all_states = set()
            for v in vendors:
                all_states.update(v['states'])
            states_str = ', '.join(sorted(all_states))

            # Summary row
            summary_data.append({
                'Platform': platform,
                'Location': location,
                'Month': month,
                'Skills': skills_str,
                'Skill_Count': len(skillset),  # Single-skill vs multi-skill
                'Vendor_Count': len(vendors),
                'States_Available': states_str
            })

            # Detail rows (one per vendor)
            for vendor in vendors:
                details_data.append({
                    'Platform': platform,
                    'Location': location,
                    'Month': month,
                    'Skills': skills_str,
                    'Vendor_ID': vendor['vendor_id'],
                    'Vendor_States': ', '.join(vendor['states']),
                    'Allocated': vendor['allocated']
                })

        # Create DataFrames
        summary_df = pd.DataFrame(summary_data)
        details_df = pd.DataFrame(details_data)

        logger.info(f"Generated bucket summary data: {len(summary_df)} buckets, {len(details_df)} vendor instances")

        return summary_df, details_df

    def _export_buckets_to_excel(self):
        """
        Export bucket structure to Excel for debugging.

        Creates two sheets:
        1. Summary: Overview of all buckets with counts
        2. Details: Full vendor details for each bucket
        """
        try:
            # Generate data using new method
            summary_df, details_df = self.generate_buckets_summary()

            # Export to Excel
            curpth = os.path.join(BASE_DIR, 'logics')
            output_path = os.path.join(curpth, 'buckets_debug.xlsx')

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='Bucket_Summary', index=False)
                details_df.to_excel(writer, sheet_name='Vendor_Details', index=False)

            logger.info(f"✓ Exported bucket structure to: {output_path}")
            logger.info(f"  - Summary sheet: {len(summary_df)} buckets")
            logger.info(f"  - Details sheet: {len(details_df)} vendor-month instances")

        except Exception as e:
            logger.warning(f"Failed to export buckets to Excel: {e}")

    def generate_buckets_after_allocation(self) -> pd.DataFrame:
        """
        Generate bucket allocation data (without Excel export).

        Returns:
            DataFrame showing allocated vs unallocated vendors per bucket
        """
        allocation_data = []

        for (platform, location, month, skillset), vendors in sorted(self.buckets.items()):
            skills_str = ' + '.join(sorted(skillset))

            # Count allocated vs unallocated
            allocated_count = sum(1 for v in vendors if v['allocated'])
            unallocated_count = sum(1 for v in vendors if not v['allocated'])

            # Get states for allocated and unallocated vendors
            allocated_states = set()
            unallocated_states = set()
            for v in vendors:
                if v['allocated']:
                    allocated_states.update(v['states'])
                else:
                    unallocated_states.update(v['states'])

            allocation_data.append({
                'Platform': platform,
                'Location': location,
                'Month': month,
                'Skills': skills_str,
                'Skill_Count': len(skillset),
                'Total_Vendors': len(vendors),
                'Allocated': allocated_count,
                'Unallocated': unallocated_count,
                'Allocation_Rate': f"{allocated_count}/{len(vendors)}" if len(vendors) > 0 else "0/0",
                'Allocated_States': ', '.join(sorted(allocated_states)) if allocated_states else '-',
                'Unallocated_States': ', '.join(sorted(unallocated_states)) if unallocated_states else '-'
            })

        allocation_df = pd.DataFrame(allocation_data)

        logger.info(f"Generated buckets after allocation data: {len(allocation_df)} buckets")
        if len(allocation_df) > 0:
            logger.info(f"  - Total allocated: {allocation_df['Allocated'].sum()}")
            logger.info(f"  - Total unallocated: {allocation_df['Unallocated'].sum()}")

        return allocation_df

    def export_buckets_after_allocation(self):
        """
        Export bucket structure AFTER allocation to show what was allocated.
        Creates a file showing allocated vs unallocated vendors.
        """
        try:
            # Generate data using new method
            allocation_df = self.generate_buckets_after_allocation()

            # Export to Excel
            curpth = os.path.join(BASE_DIR, 'logics')
            output_path = os.path.join(curpth, 'buckets_after_allocation.xlsx')

            allocation_df.to_excel(output_path, index=False, engine='openpyxl')

            logger.info(f"✓ Exported post-allocation buckets to: {output_path}")
            logger.info(f"  - Total buckets: {len(allocation_df)}")

        except Exception as e:
            logger.warning(f"Failed to export post-allocation buckets: {e}")

    def allocate(self, platform: str, state: str, month: str, worktype: str, fte_required: float) -> tuple:
        """
        Allocate resources for a demand request.

        CRITICAL: Each vendor can only be allocated ONCE (no double-counting).
        Vendors with StateList=[FL, GA, N/A] can fulfill FL, GA, or N/A demands,
        but once allocated, they're marked and cannot be reused.

        Priority:
        1. Single-skill exact match (e.g., only "FTC-Basic/Non MMP")
        2. Multi-skill containing the worktype (e.g., "FTC-Basic/Non MMP" + "ADJ")

        Args:
            platform: Platform name (e.g., "Amisys Medicaid Domestic")
                     Special case: "OIC Volumes" - locality extracted from worktype
            state: State code (e.g., "MI") or "N/A" for any state
            month: Month name (e.g., "March")
            worktype: Exact worktype from demand (e.g., "FTC-Basic/Non MMP")
                     Special case for OIC Volumes: Contains "domestic" → Domestic, else → Global
            fte_required: Number of FTEs needed

        Returns:
            tuple: (allocated_amount, shortage_amount)
        """
        # Skip allocation if no FTE required
        if fte_required <= 0:
            return 0, 0

        # Parse LOB to extract platform and locality
        parsed_lob = parse_main_lob(platform)
        lob_platform = parsed_lob.get('platform', platform)
        lob_locality = parsed_lob.get('locality', '')

        # SPECIAL CASE: OIC Volumes - locality is in worktype column
        # Check if market contains "OIC Volumes"
        is_oic_volumes = 'oic' in str(platform).lower() and 'volumes' in str(platform).lower()

        if is_oic_volumes:
            # Check if worktype contains "domestic" (case-insensitive)
            worktype_lower = str(worktype).lower()
            if 'domestic' in worktype_lower:
                lob_locality = 'Domestic'
                logger.debug(f"[SPECIAL CASE] OIC Volumes: Found 'domestic' in worktype '{worktype}' → locality = Domestic")
            else:
                # Default to Global if domestic not found
                lob_locality = 'Global'
                logger.debug(f"[SPECIAL CASE] OIC Volumes: 'domestic' not found in worktype '{worktype}' → locality = Global")

        # Normalize inputs
        platform_normalized = str(lob_platform).strip().split()[0].upper() if lob_platform and str(lob_platform).lower() != 'nan' else lob_platform
        location_normalized = normalize_locality(lob_locality)
        state_normalized = str(state).strip().upper()
        month_normalized = str(month).strip().title()
        worktype_normalized = self._normalize_text(worktype).lower()

        allocated = 0
        remaining = fte_required

        # Debug logging for first few allocations
        if len(self.allocation_history) < 5:
            logger.info(f"ALLOCATE REQUEST: platform={platform!r} → parsed={parsed_lob}, platform_normalized={platform_normalized!r}, location_normalized={location_normalized!r}, state={state!r} → {state_normalized!r}, month={month!r} → {month_normalized!r}, worktype={worktype!r} → {worktype_normalized!r}, fte_required={fte_required}")

        # Priority 1: Single-skill exact match
        single_skillset = frozenset({worktype_normalized})
        allocated_from_single = self._allocate_from_skillset(
            platform_normalized, location_normalized, month_normalized, single_skillset,
            state_normalized, remaining, "single-skill",
            platform, state, worktype
        )
        allocated += allocated_from_single
        remaining -= allocated_from_single

        # Priority 2: Multi-skill containing this worktype
        if remaining > 0:
            if len(self.allocation_history) < 5:
                logger.info(f"  Priority 2: Looking for multi-skill containing '{worktype_normalized}'")

            # Find all buckets with multi-skills containing this worktype
            for bucket_key, vendors in self.buckets.items():
                if remaining <= 0:
                    break

                plat, loc, mon, skillset = bucket_key
                if plat == platform_normalized and loc == location_normalized and mon == month_normalized:
                    if len(skillset) > 1 and worktype_normalized in skillset:
                        allocated_from_multi = self._allocate_from_vendor_list(
                            vendors, state_normalized, remaining, bucket_key,
                            platform, state, month, worktype
                        )
                        allocated += allocated_from_multi
                        remaining -= allocated_from_multi

        # Track history
        self.allocation_history.append({
            'platform': platform,
            'state': state,
            'month': month,
            'worktype': worktype,
            'requested': fte_required,
            'allocated': allocated,
            'shortage': remaining
        })

        # Final debug log
        if len(self.allocation_history) <= 5:
            logger.info(f"  RESULT: Allocated {allocated}/{fte_required}, Shortage: {remaining}")

        return allocated, remaining

    def _allocate_from_skillset(self, platform: str, location: str, month: str, skillset: frozenset,
                                 demand_state: str, fte_required: float, label: str,
                                 original_platform: str, original_state: str, worktype: str) -> float:
        """
        Allocate resources from a specific skillset bucket.

        Args:
            platform: Normalized platform
            location: Normalized location (Domestic/Global)
            month: Normalized month
            skillset: Skillset to search for
            demand_state: State required by demand
            fte_required: FTEs needed
            label: Debug label
            original_platform: Original platform from demand (for tracking)
            original_state: Original state from demand (for tracking)
            worktype: Worktype from demand (for tracking)

        Returns:
            float: Amount allocated
        """
        bucket_key = (platform, location, month, skillset)

        if bucket_key not in self.buckets:
            return 0.0

        if len(self.allocation_history) < 5:
            logger.info(f"  Priority 1: Looking for {label} match {skillset} (location={location})")

        vendors = self.buckets[bucket_key]
        allocated = self._allocate_from_vendor_list(
            vendors, demand_state, fte_required, bucket_key,
            original_platform, original_state, month, worktype
        )

        return allocated

    def _allocate_from_vendor_list(self, vendors: list, demand_state: str,
                                     fte_required: float, bucket_key: tuple,
                                     platform: str, state: str, month: str, worktype: str) -> float:
        """
        Allocate from a list of vendors, checking state compatibility.

        Args:
            vendors: List of vendor dicts with 'states' and 'allocated' flag
            demand_state: State required by demand (can be N/A)
            fte_required: FTEs needed
            bucket_key: For debug logging
            platform: Platform from demand (for tracking)
            state: State from demand (for tracking)
            month: Month from demand (for tracking)
            worktype: Worktype from demand (for tracking)

        Returns:
            float: Amount allocated
        """
        allocated = 0.0

        for vendor in vendors:
            if allocated >= fte_required:
                break

            # Skip if already allocated
            if vendor['allocated']:
                continue

            # Check if vendor can work in this state
            # Note: Every vendor has 'N/A' in their StateList, so N/A demands always match
            if demand_state in vendor['states']:
                # Allocate this vendor (1 FTE)
                vendor['allocated'] = True

                # Store allocation details in vendor record
                allocation_details = {
                    'platform': platform,
                    'state': state,
                    'month': month,
                    'worktype': worktype
                }
                vendor['allocation_details'] = allocation_details

                # Add to reverse lookup index for O(1) report generation
                vendor_df_idx = vendor['vendor_df_idx']
                if vendor_df_idx not in self.vendor_allocations:
                    self.vendor_allocations[vendor_df_idx] = {}
                self.vendor_allocations[vendor_df_idx][month] = allocation_details

                allocated += 1

                if len(self.allocation_history) < 5:
                    logger.info(f"    ✓ Allocated 1 FTE (vendor_id={vendor['vendor_id']}, states={vendor['states']}) from {bucket_key}")

        return allocated

    def get_summary_report(self) -> dict:
        """
        Generate summary report with initial/allocated/unutilized FTE counts.

        Returns:
            dict with keys: summary, by_category
        """
        # Calculate totals (count vendors, not duplicates)
        total_initial = sum(len(vendors) for vendors in self.initial_state.values())

        # Count allocated vendors
        total_allocated = sum(
            sum(1 for v in vendors if v['allocated'])
            for vendors in self.buckets.values()
        )

        total_current = total_initial - total_allocated

        # Calculate by category (based on skillset length)
        single_skill_initial = sum(
            len(vendors) for (plat, loc, mon, skillset), vendors in self.initial_state.items()
            if len(skillset) == 1
        )
        single_skill_allocated = sum(
            sum(1 for v in vendors if v['allocated'])
            for (plat, loc, mon, skillset), vendors in self.buckets.items()
            if len(skillset) == 1
        )

        multi_skill_initial = sum(
            len(vendors) for (plat, loc, mon, skillset), vendors in self.initial_state.items()
            if len(skillset) > 1
        )
        multi_skill_allocated = sum(
            sum(1 for v in vendors if v['allocated'])
            for (plat, loc, mon, skillset), vendors in self.buckets.items()
            if len(skillset) > 1
        )

        total_requested = sum(h['requested'] for h in self.allocation_history)

        return {
            'summary': {
                'total_initial_fte': total_initial,
                'total_allocated_fte': total_allocated,
                'total_unutilized_fte': total_current,
                'allocation_success_rate': total_allocated / total_requested if total_requested > 0 else 0
            },
            'by_category': {
                'single_skill': {
                    'initial': single_skill_initial,
                    'allocated': single_skill_allocated,
                    'remaining': single_skill_initial - single_skill_allocated
                },
                'multi_skill': {
                    'initial': multi_skill_initial,
                    'allocated': multi_skill_allocated,
                    'remaining': multi_skill_initial - multi_skill_allocated
                }
            }
        }

    def get_unmet_demand_report(self) -> pd.DataFrame:
        """
        Generate report of all allocation requests with shortages.

        Returns:
            DataFrame with columns: Platform, State, Month, Worktype, Requested, Allocated, Shortage
        """
        # Filter history to only show entries with shortages
        shortages = [h for h in self.allocation_history if h['shortage'] > 0]

        if not shortages:
            return pd.DataFrame(columns=['Platform', 'State', 'Month', 'Worktype', 'Requested', 'Allocated', 'Shortage'])

        return pd.DataFrame(shortages).rename(columns={
            'platform': 'Platform',
            'state': 'State',
            'month': 'Month',
            'worktype': 'Worktype',
            'requested': 'Requested',
            'allocated': 'Allocated',
            'shortage': 'Shortage'
        })

    def get_unutilized_report(self, output_df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate report of unutilized vendor resources.

        IMPORTANT: Only shows unutilized resources for worktypes that appeared in demand.

        Args:
            output_df: The demand DataFrame to filter relevant worktypes

        Returns:
            DataFrame with columns: Platform, Month, Skills, States, Count
        """
        # Get set of demanded worktypes (normalized)
        demanded_worktypes = {
            self._normalize_text(wt).lower()
            for wt in output_df[('Centene Capacity plan', 'Case type')].unique()
            if wt and str(wt).lower() not in {'nan', 'none', ''}
        }

        unutilized = []
        for (platform, location, month, skillset), vendors in self.buckets.items():
            # Count unallocated vendors
            unallocated_vendors = [v for v in vendors if not v['allocated']]

            if len(unallocated_vendors) > 0:
                # Check if ANY skill in this skillset was demanded
                if skillset & demanded_worktypes:  # Set intersection
                    skills_str = ' + '.join(sorted(skillset))

                    # Collect unique states from unallocated vendors
                    all_states = set()
                    for v in unallocated_vendors:
                        all_states.update(v['states'])

                    unutilized.append({
                        'Platform': platform,
                        'Location': location,
                        'Month': month,
                        'Skills': skills_str,
                        'States': ', '.join(sorted(all_states)),
                        'Count': len(unallocated_vendors)
                    })

        if not unutilized:
            return pd.DataFrame(columns=['Platform', 'Location', 'Month', 'Skills', 'States', 'Count'])

        return pd.DataFrame(unutilized).sort_values(['Platform', 'Location', 'Month', 'Skills'])

    def generate_roster_allotment(self) -> pd.DataFrame:
        """
        Generate vendor-level allocation data (without Excel export).

        Returns:
            DataFrame with one row per vendor, showing:
            - Vendor identification (FirstName, LastName, CN)
            - Work details (PrimaryPlatform, NewWorkType, Location, State)
            - Allocation status (Allocated/Not Allocated)
            - Per-month allocation details (LOB, State, Worktype)

        Uses O(n×m) optimized lookup via self.vendor_allocations reverse index.
        """
        logger.info("Generating roster allotment data...")

        report_data = []

        # Iterate through all vendors in original filtered vendor_df
        for idx, vendor_row in self.vendor_df_original.iterrows():
            # Extract vendor identification
            first_name = vendor_row.get('FirstName', '')
            last_name = vendor_row.get('LastName', '')
            cn = vendor_row.get('CN', '')

            # Extract work details
            primary_platform = vendor_row.get('PrimaryPlatform', '')
            new_worktype = vendor_row.get('NewWorkType', '')
            location = vendor_row.get('Location', '')
            state = vendor_row.get('State', '')  # Original state without normalization

            # Check if this vendor was allocated (in any month)
            vendor_allocations = self.vendor_allocations.get(idx, {})
            is_allocated = len(vendor_allocations) > 0
            status = "Allocated" if is_allocated else "Not Allocated"

            # Build row data
            row_data = {
                'FirstName': first_name,
                'LastName': last_name,
                'CN': cn,
                'PrimaryPlatform': primary_platform,
                'NewWorkType': new_worktype,
                'Location': location,
                'State': state,  # Original vendor state
                'Status': status
            }

            # Add per-month allocation details
            for month in self.month_headers:
                allocation = vendor_allocations.get(month)

                if allocation:
                    # Vendor was allocated in this month
                    row_data[f'{month}_LOB'] = allocation['platform']
                    row_data[f'{month}_State'] = allocation['state']
                    row_data[f'{month}_Worktype'] = allocation['worktype']
                else:
                    # Vendor not allocated in this month
                    row_data[f'{month}_LOB'] = 'Not Allocated'
                    row_data[f'{month}_State'] = '-'
                    row_data[f'{month}_Worktype'] = '-'

            report_data.append(row_data)

        # Create DataFrame
        report_df = pd.DataFrame(report_data)

        logger.info(f"Generated roster allotment data: {len(report_df)} vendors")
        if len(report_df) > 0:
            logger.info(f"  - Allocated vendors: {(report_df['Status'] == 'Allocated').sum()}")
            logger.info(f"  - Unallocated vendors: {(report_df['Status'] == 'Not Allocated').sum()}")

        return report_df

    def export_roster_allotment_report(self):
        """
        Export vendor-level allocation report showing allocation status for each vendor.

        Creates roster_allotment.xlsx with one row per vendor, showing:
        - Vendor identification (FirstName, LastName, CN)
        - Work details (PrimaryPlatform, NewWorkType, Location, State)
        - Allocation status (Allocated/Not Allocated)
        - Per-month allocation details (LOB, State, Worktype)

        Uses O(n×m) optimized lookup via self.vendor_allocations reverse index.
        """
        try:
            # Generate data using new method
            report_df = self.generate_roster_allotment()

            # Export to Excel
            curpth = os.path.join(BASE_DIR, 'logics')
            output_path = os.path.join(curpth, 'roster_allotment.xlsx')

            report_df.to_excel(output_path, index=False, engine='openpyxl')

            logger.info(f"✓ Exported roster allotment report to: {output_path}")
            logger.info(f"  - Total vendors: {len(report_df)}")

        except Exception as e:
            logger.warning(f"Failed to export roster allotment report: {e}")


def get_capacity(row, month, calculations: Calculations):
    """
    Calculate Capacity using work-type-specific configuration parameters.

    Determines work type (Domestic/Global) from LOB parsing, with special handling
    for OIC Volumes where work type is extracted from the Case Type column.

    Formula: Target_CPH * FTE_Available * (1 - Shrinkage) * WorkingDays * WorkHours
    """
    try:
        target_cph = row[('Centene Capacity plan', 'Target CPH')]
        fte_available = row[('FTE Avail', month)]
        main_lob = row[('Centene Capacity plan', 'Main LOB')]
        case_type = row[('Centene Capacity plan', 'Case type')]

        # Determine work type based on LOB parsing
        parsed_lob = parse_main_lob(main_lob)
        lob_locality = parsed_lob.get('locality', '')

        # SPECIAL CASE: OIC Volumes - locality determined from worktype column
        is_oic_volumes = 'oic' in str(main_lob).lower() and 'volumes' in str(main_lob).lower()
        if is_oic_volumes:
            case_type_lower = str(case_type).lower()
            work_type = 'Domestic' if 'domestic' in case_type_lower else 'Global'
        else:
            # Normalize locality to Domestic/Global
            work_type = 'Domestic' if 'domestic' in str(lob_locality).lower() else 'Global'

        # Get work-type-specific configuration
        year = calculations.data_year if calculations.data_year else datetime.now().year  # Fallback to current year
        config = calculations.get_config_for_worktype(month, year, work_type)

        no_of_days = config['working_days']
        shrinkage = config['shrinkage']
        workhours = config['work_hours']

        logging.debug(f"FTE Avail for {month}: {fte_available}, work_type: {work_type}")
        capacity = target_cph * fte_available * (1 - shrinkage) * no_of_days * workhours
        return capacity

    except (KeyError, TypeError) as e:
        logging.error(f"Error in get_capacity for {month}: {e}", exc_info=True)
        return 0
    except ValueError:
        # ValueError indicates missing month config - must propagate to halt execution
        raise

# Legacy per-file vendor filtering - no longer needed with ResourceAllocator
# (ResourceAllocator handles platform/state matching internally)
# def filter_vendor_df(file_name, vendor_df):
#     platform = file_name.split(" ")[0]
#     location = 'Domestic' if 'domestic' in file_name.lower() else 'Global'
#     filtered_df = vendor_df[
#         (vendor_df['PartofProduction'].isin(['Production', 'Ramp'])) &
#         (vendor_df['Location'].str.lower() == location.lower()) &
#         (vendor_df['BeelineTitle'] == 'Claims Analyst') &
#         (vendor_df['PrimaryPlatform'].str.lower() == platform.lower())
#     ]
#     return filtered_df

def initialize_output_excel(month_headers:List[str]):
    wb = Workbook()
    ws = wb.active
    logger.debug(f"input month headers: {month_headers}")
    ws.title = "Capacity plan"
    capacity_plan_headers = ["Main LOB", "State", "Case type", "Call Type ID", "Target CPH"]
    columns = capacity_plan_headers + month_headers * 4
    ws.merge_cells("A1:E1")
    ws.merge_cells("F1:K1")
    ws.merge_cells("L1:Q1")
    ws.merge_cells("R1:W1")
    ws.merge_cells("X1:AC1")
    ws["A1"] = "Centene Capacity plan"
    ws["F1"] = "Client Forecast"
    ws["L1"] = "FTE Required"
    ws["R1"] = "FTE Avail"
    ws["X1"] = "Capacity"
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 32):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = header_alignment
    for col, header in enumerate(columns, start=1):
        ws.cell(row=2, column=col, value=header).font = header_font
        ws.cell(row=2, column=col).alignment = header_alignment
    ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=len(columns)).coordinate}"
    wb.save(output_file)

def get_columns_between_column_names(
    df: pd.DataFrame,
    col_level: int = 0,
    start_col_name: str = None,
    end_col_name: str = None
) -> List[str]:
    """
    Returns a list of column names (from the specified row of columns) between start_col_name and the first end_col_name, without duplicates.
    Edge cases:
      - If both start_col_name and end_col_name are missing, returns all columns.
      - If only start_col_name is present, returns columns from that name to the end.
      - If only end_col_name is present, returns columns from start up to (not including) that name.
      - If both are present, returns columns between them.
    """
    if df is None or df.empty or not hasattr(df, "columns"):
        return []

    first_row = df.columns.get_level_values(col_level)
    first_row_str = [str(col).strip() for col in first_row]
    first_row_lower = [col.lower() for col in first_row_str]

    start_idx = None
    end_idx = None

    if start_col_name:
        start_col_name_lower = str(start_col_name).strip().lower()
        if start_col_name_lower in first_row_lower:
            start_idx = first_row_lower.index(start_col_name_lower)
    if end_col_name:
        end_col_name_lower = str(end_col_name).strip().lower()
        # Find first occurrence of end_col_name after start_idx
        search_start = (start_idx + 1) if start_idx is not None else 0
        try:
            end_idx = first_row_lower.index(end_col_name_lower, search_start)
        except ValueError:
            end_idx = None

    # Logic for edge cases
    if start_idx is not None and end_idx is not None and end_idx > start_idx:
        cols = first_row_str[start_idx + 1:end_idx]
    elif start_idx is not None:
        cols = first_row_str[start_idx + 1:]
    elif end_idx is not None:
        cols = first_row_str[:end_idx]
    else:
        cols = first_row_str

    # Remove duplicates while preserving order
    seen = set()
    unique_cols = [x for x in cols if x and not (x in seen or seen.add(x))]
    return unique_cols

def get_nonmmp_columns(df:pd.DataFrame) -> List[str]:
    """
    Returns a list of unique third-level column names
    from a multi-indexed DataFrame, where:
    - First-level column contains 'Forecast - Volume'
    - Third-level column does not contain 'Total'
    """
    # Ensure we are working with MultiIndex columns
    if not isinstance(df.columns, pd.MultiIndex):
        raise ValueError("DataFrame must have MultiIndex columns (3 levels).")

    # Filter columns where first level contains 'Forecast - Volume'
    filtered = [col for col in df.columns if "Forecast - Volume" in str(col[0])]

    # Extract the third-level names
    third_level_names = [col[2] for col in filtered if "Total" not in str(col[2])]

    # Return unique names as a list
    return sorted(set(third_level_names))


def get_roster_file_metadata(requested_month: str, requested_year: int, core_utils) -> Dict:
    """
    Get metadata about which roster file was actually used for allocation.

    Detects if fallback to latest roster data occurred when requested month/year
    is not available. This is critical for audit trail to understand which data
    was used in allocation calculations.

    Args:
        requested_month: Month requested for roster data (e.g., "January")
        requested_year: Year requested for roster data (e.g., 2025)
        core_utils: CoreUtils instance for database access

    Returns:
        Dictionary with keys:
            - filename: Original uploaded filename from database
            - month: Actual month of roster data used
            - year: Actual year of roster data used
            - was_fallback: True if latest data was used instead of requested
    """
    db_manager = core_utils.get_db_manager(ProdTeamRosterModel, limit=1, skip=0, select_columns=None)

    # Try to get requested month/year
    with db_manager.SessionLocal() as session:
        requested_record = session.query(ProdTeamRosterModel).filter(
            ProdTeamRosterModel.Month == requested_month,
            ProdTeamRosterModel.Year == requested_year
        ).first()

        if requested_record:
            # Found exact match
            logger.info(f"Using requested roster: {requested_month} {requested_year} - {requested_record.UploadedFile}")
            return {
                'filename': requested_record.UploadedFile,
                'month': requested_month,
                'year': requested_year,
                'was_fallback': False
            }

        # Not found - get latest (fallback)
        logger.warning(f"Roster for {requested_month} {requested_year} not found, falling back to latest")
        latest_record = session.query(ProdTeamRosterModel).order_by(
            ProdTeamRosterModel.Year.desc(),
            ProdTeamRosterModel.CreatedDateTime.desc()
        ).first()

        if latest_record:
            logger.warning(
                f"Using latest roster: {latest_record.Month} {latest_record.Year} - {latest_record.UploadedFile}"
            )
            return {
                'filename': latest_record.UploadedFile,
                'month': latest_record.Month,
                'year': latest_record.Year,
                'was_fallback': True
            }

        # No roster data at all
        logger.error("No roster data found in database")
        return {
            'filename': 'No roster data found',
            'month': 'N/A',
            'year': 0,
            'was_fallback': True
        }


def build_config_snapshot(calculations: Calculations) -> Dict:
    """
    Build configuration snapshot from loaded configs in calculations cache.

    Captures all month configurations that were actually used during this execution
    for audit trail and debugging purposes.

    Args:
        calculations: Calculations instance with populated _config_cache

    Returns:
        Dictionary with structure:
        {
            "month_config": {
                "January 2025": {
                    "Domestic": {...},
                    "Global": {...}
                },
                "February 2025": {
                    "Domestic": {...},
                    "Global": {...}
                }
            }
        }
    """
    config_snapshot = {"month_config": {}}

    # Extract configs from the calculations cache
    for (month, year, work_type), config in calculations._config_cache.items():
        # Create month-year key for grouping
        month_year_key = f"{month} {year}"

        # Initialize month-year entry if not exists
        if month_year_key not in config_snapshot["month_config"]:
            config_snapshot["month_config"][month_year_key] = {}

        # Store config under work_type within the month-year
        config_snapshot["month_config"][month_year_key][work_type] = {
            "working_days": config['working_days'],
            "occupancy": config['occupancy'],
            "shrinkage": config['shrinkage'],
            "work_hours": config['work_hours']
        }

    return config_snapshot


def process_files(data_month: str, data_year: int, forecast_file_uploaded_by: str, forecast_filename: str):
    """
    Simulates logic after forecast file upload and updates the processed forecast data.

    Includes comprehensive execution tracking with audit trail of source files,
    configuration snapshots, and error details.
    """
    execution_id = None
    start_time = datetime.now()

    try:
        logging.info("Starting file processing")

        # STEP 1: Capture roster file metadata BEFORE loading data
        # This detects fallback scenarios for audit trail
        roster_metadata = get_roster_file_metadata(data_month, data_year, core_utils)
        roster_filename = roster_metadata['filename']
        roster_month_used = roster_metadata['month']
        roster_year_used = roster_metadata['year']
        roster_was_fallback = roster_metadata['was_fallback']

        # STEP 2: Start execution tracking
        execution_id = start_execution(
            month=data_month,
            year=data_year,
            forecast_filename=forecast_filename,
            roster_filename=roster_filename,
            roster_month_used=roster_month_used,
            roster_year_used=roster_year_used,
            roster_was_fallback=roster_was_fallback,
            uploaded_by=forecast_file_uploaded_by
        )

        # STEP 3: Update status to IN_PROGRESS
        update_status(execution_id, 'IN_PROGRESS')
        logging.info(f"Execution tracking started: {execution_id}")

        # Load and filter vendor data upfront
        vendor_df_raw = get_latest_or_requested_dataframe('prod_team_roster', data_month, data_year)

        # Filter to eligible vendors only (Production/Ramp, Claims Analyst)
        logging.info(f"Raw vendor data: {vendor_df_raw.shape}")
        vendor_df = vendor_df_raw[
            (vendor_df_raw['PartofProduction'].isin(['Production', 'Ramp'])) &
            (vendor_df_raw['BeelineTitle'] == 'Claims Analyst')
        ].copy()
        logging.info(f"Filtered vendor data: {vendor_df.shape} (eligible vendors only)")

        # Clean column names
        vendor_df.columns = vendor_df.columns.str.replace("\n", "").str.strip()

        # Verify required columns exist
        required_cols = ['PrimaryPlatform', 'State', 'NewWorkType']
        missing_cols = [col for col in required_cols if col not in vendor_df.columns]
        if missing_cols:
            logging.error(f"Missing required columns in vendor_df: {missing_cols}")
            logging.error(f"Available columns: {list(vendor_df.columns)}")
            raise ValueError(f"Vendor DataFrame missing required columns: {missing_cols}")

        logging.info(f"Vendor data sample - Platform: {vendor_df['PrimaryPlatform'].unique()[:5]}")
        logging.info(f"Vendor data sample - States: {vendor_df['State'].unique()[:10]}")
        logging.info(f"Vendor data sample - NewWorkType (first 5): {vendor_df['NewWorkType'].head().tolist()}")

        month_headers = get_forecast_months_list(data_month, data_year, forecast_filename)
        calculations = Calculations(data_month=data_month, data_year=data_year)
        initialize_output_excel(month_headers)
        output_dfs = []
        file_types = get_all_model_dataframes_dict(data_month, data_year)

        for file_type, directory in file_types.items():
            for file_name, df in directory.items():
                logger.info(f"Processing {file_type} file: {file_name}")
                client_names, states, work_types = [], [], []
                # logger.debug(f"Column values: ")
                # for col in df.columns:
                #     logger.debug(f"column value ---- {col}")
                if file_type == 'medicare_medicaid_mmp':
                    try:
                        work_types = get_columns_between_column_names(df,0,'Mo St', 'Year')
                        logging.info(f"Extracted work_types: {work_types}, length: {len(work_types)}, top 5 worktypes: {work_types[:5]}")
                        states = list(set(s for s in df[('State', 'State')].dropna() if s != 0 and isinstance(s, str)))
                        logging.info(f"Extracted states: {states}, length: {len(states)}")
                        if not work_types or not states:
                            logging.warning(f"No valid work types or states for {file_name}. Skipping.")
                            continue
                        total_rows = len(states) * len(work_types)
                        states_expanded = [s for s in states for _ in work_types]
                        work_types_expanded = work_types * len(states)
                        client_names = [file_name] * total_rows
                        logging.info(f"Expected rows: {total_rows}, states: {len(states_expanded)}, work_types: {len(work_types_expanded)}, client_names: {len(client_names)}")
                    except Exception as e:
                        logging.error(f"Error reading MMP file {file_name}: {e}")
                        continue

                elif file_type == 'medicare_medicaid_nonmmp':
                    try:
                        df.columns = df.columns.map(lambda x: tuple(i if 'Unnamed' not in str(i) else '' for i in x))

                        work_types = get_nonmmp_columns(df)
                        logging.info(f"Extracted work_types: {work_types}, length: {len(work_types)}")
                        states = list(set(df[('', '', 'State')].dropna()))
                        logging.info(f"Extracted states: {states}, length: {len(states)}")
                        total_rows = len(states) * len(work_types)
                        states_expanded = [s for s in states for _ in work_types]
                        work_types_expanded = work_types * len(states)
                        client_names = [file_name] * total_rows
                        logging.info(f"Expected rows: {total_rows}, states: {len(states_expanded)}, work_types: {len(work_types_expanded)}, client_names: {len(client_names)}")
                    except Exception as e:
                        logging.error(f"Error reading non-MMP file {file_name}: {e}")
                        continue

                elif file_type == 'medicare_medicaid_summary' and 'mmp' not in file_name.lower():
                    try:
                        if df.empty:
                            logging.warning(f"Empty DataFrame for {file_name}. Skipping.")
                            continue
                        logging.info(f"DataFrame shape for {file_name}: {df.shape}")
                        df.columns = df.columns.map(lambda x: tuple(i if 'Unnamed' not in str(i) else '' for i in x))
                        # df.columns = df.columns.sort_values()
                        first_col = file_name.split("-summary")[0]
                        logging.info(f"Summary file columns: {df.columns.tolist()}")
                        work_type_col = None
                        for col in df.columns:
                            if col[2].lower() in ['work type', 'worktype', 'case type']:
                                work_type_col = col
                                break
                        if work_type_col:
                            try:
                                work_type_values = list(df[work_type_col].dropna())
                                logging.info(f"Work Type column values: {work_type_values}")
                            except Exception as e:
                                logging.warning(f"Failed to log Work Type values for {file_name}: {e}")
                                work_type_values = []
                            work_types = [str(wkt) for wkt in df[work_type_col].fillna('') if str(wkt).lower() != "total" and str(wkt).strip()]
                        else:
                            logging.warning(f"No 'Work Type' column found in {file_name}. Available columns: {df.columns.tolist()}")
                            continue
                        logging.info(f"Extracted work_types: {work_types}, length: {len(work_types)}")
                        states = ["N/A"] * len(work_types)
                        client_names = [first_col] * len(work_types)
                        total_rows = len(work_types)
                        states_expanded = states
                        work_types_expanded = work_types
                        logging.info(f"Expected rows: {total_rows}, states: {len(states_expanded)}, work_types: {len(work_types_expanded)}, client_names: {len(client_names)}")
                    except Exception as e:
                        logging.error(f"Error reading summary file {file_name}: {e}")
                        continue

                if not work_types_expanded:
                    logging.warning(f"No valid work types found for {file_name}. Skipping.")
                    continue

                # Initialize output_df with correct number of rows
                column_template = pd.read_excel(output_file, header=[0, 1]).columns
                output_df = pd.DataFrame(index=range(total_rows), columns=column_template)
                try:
                    output_df[('Centene Capacity plan', 'Main LOB')] = client_names
                    output_df[('Centene Capacity plan', 'State')] = states_expanded
                    output_df[('Centene Capacity plan', 'Case type')] = work_types_expanded
                    output_df[('Centene Capacity plan', 'temp_Case type')] = output_df[('Centene Capacity plan', 'Case type')].apply(get_temp_casetype)
                    output_df[('Centene Capacity plan', 'Call Type ID')] = output_df[('Centene Capacity plan', 'Main LOB')].astype(str)+ " " + output_df[('Centene Capacity plan', 'temp_Case type')].astype(str)
                    output_df.drop(columns=[('Centene Capacity plan', 'temp_Case type')], inplace=True)
                except ValueError as e:
                    logging.error(f"Length mismatch in {file_name}: {e}")
                    logging.info(f"client_names: {len(client_names)}, states: {len(states_expanded)}, work_types: {len(work_types_expanded)}")
                    continue
                for month in month_headers:
                    if file_type == 'medicare_medicaid_mmp':
                        output_df[('Client Forecast', month)] = output_df.apply(lambda row: get_value(row, month, 'medicare_medicaid_mmp', df), axis=1)
                    elif file_type == 'medicare_medicaid_nonmmp':
                        output_df[('Client Forecast', month)] = output_df.apply(lambda row: get_value(row, month, 'medicare_medicaid_nonmmp', df), axis=1)
                    elif file_type == 'medicare_medicaid_summary':
                        output_df[('Client Forecast', month)] = output_df.apply(lambda row: get_value(row, month, 'medicare_medicaid_summary', df), axis=1)

                output_df[('Centene Capacity plan', 'Target CPH')] = output_df.apply(calculations.get_target_cph, axis=1)
                for month in month_headers:
                    output_df[('FTE Required', month)] = output_df.apply(lambda row: get_fte_required(row, month, calculations), axis=1)
                    output_df[('FTE Required', month)] = output_df[('FTE Required', month)].apply(lambda x: 0.5 if 0 < x < 0.5 else x)
                output_df = output_df.fillna(0).infer_objects(copy=False)
                output_df[output_df.select_dtypes(include=['number']).columns] = output_df.select_dtypes(include=['number']).round().astype(int)

                # Initialize FTE Avail columns (will be populated after consolidation)
                for month in month_headers:
                    output_df[('FTE Avail', month)] = 0

                output_dfs.append(output_df)

        if output_dfs:
            logging.info(f"Collected {len(output_dfs)} file outputs, consolidating...")
            consolidated_df = pd.concat(output_dfs, ignore_index=True)
            logging.info(f"Consolidated DataFrame shape: {consolidated_df.shape}")

            # Debug: Show sample demand data
            logging.info("=== DEMAND DATA SAMPLE ===")
            logging.info(f"Unique platforms in demand: {consolidated_df[('Centene Capacity plan', 'Main LOB')].unique()[:10]}")
            logging.info(f"Unique states in demand: {consolidated_df[('Centene Capacity plan', 'State')].unique()[:10]}")
            logging.info(f"Unique worktypes in demand (first 10): {consolidated_df[('Centene Capacity plan', 'Case type')].unique()[:10]}")

            # Initialize ResourceAllocator with complete demand data
            logging.info("=== INITIALIZING RESOURCE ALLOCATOR ===")
            allocator = ResourceAllocator(vendor_df, consolidated_df, month_headers)

            # Apply allocation row by row
            logging.info("Starting resource allocation...")
            allocation_start = datetime.now()
            for idx, row in consolidated_df.iterrows():
                platform = row[('Centene Capacity plan', 'Main LOB')]
                state = row[('Centene Capacity plan', 'State')]
                worktype = row[('Centene Capacity plan', 'Case type')]

                for month in month_headers:
                    fte_required = row[('FTE Required', month)]
                    allocated, _ = allocator.allocate(platform, state, month, worktype, fte_required)
                    consolidated_df.at[idx, ('FTE Avail', month)] = allocated

            allocation_end = datetime.now()
            logging.info(f"Allocation completed in {allocation_end - allocation_start}")

            # Export buckets after allocation for debugging (Excel files)
            logging.info("Exporting post-allocation bucket state...")
            allocator.export_buckets_after_allocation()

            # Export roster allotment report (Excel file)
            logging.info("Exporting roster allotment report...")
            allocator.export_roster_allotment_report()

            # Save allocation reports to database
            logging.info("Saving allocation reports to database...")
            try:
                # Get DBManager for allocation reports using core_utils
                db_manager = core_utils.get_db_manager(
                    AllocationReportsModel,
                    limit=1000,
                    skip=0,
                    select_columns=None
                )

                # Generate and save bucket summary report (with details)
                summary_df, details_df = allocator.generate_buckets_summary()
                # Combine summary and details into single report for storage
                # Add a 'Type' column to distinguish summary from details
                summary_df['ReportSection'] = 'Summary'
                details_df['ReportSection'] = 'Details'
                bucket_summary_combined = pd.concat([summary_df, details_df], ignore_index=True)

                db_manager.save_allocation_report(
                    df=bucket_summary_combined,
                    execution_id=execution_id,
                    month=data_month,
                    year=data_year,
                    report_type='bucket_summary',
                    created_by=forecast_file_uploaded_by,
                    updated_by=forecast_file_uploaded_by
                )
                logging.info(f"✓ Saved bucket_summary report to database (execution_id: {execution_id})")

                # Generate and save buckets after allocation report
                buckets_after_df = allocator.generate_buckets_after_allocation()
                db_manager.save_allocation_report(
                    df=buckets_after_df,
                    execution_id=execution_id,
                    month=data_month,
                    year=data_year,
                    report_type='bucket_after_allocation',
                    created_by=forecast_file_uploaded_by,
                    updated_by=forecast_file_uploaded_by
                )
                logging.info(f"✓ Saved bucket_after_allocation report to database (execution_id: {execution_id})")

                # Generate and save roster allotment report
                roster_allotment_df = allocator.generate_roster_allotment()
                db_manager.save_allocation_report(
                    df=roster_allotment_df,
                    execution_id=execution_id,
                    month=data_month,
                    year=data_year,
                    report_type='roster_allotment',
                    created_by=forecast_file_uploaded_by,
                    updated_by=forecast_file_uploaded_by
                )
                logging.info(f"✓ Saved roster_allotment report to database (execution_id: {execution_id})")

            except Exception as e:
                logging.error(f"Failed to save allocation reports to database: {e}")
                logging.warning("Continuing with forecast processing despite database save failure...")

            # Calculate capacity
            for month in month_headers:
                consolidated_df[('Capacity', month)] = consolidated_df.apply(lambda row: get_capacity(row, month, calculations), axis=1)

            # Final cleanup
            consolidated_df = consolidated_df.fillna(0).infer_objects(copy=False)
            consolidated_df[consolidated_df.select_dtypes(include=['number']).columns] = consolidated_df.select_dtypes(include=['number']).round().astype(int)

            # Generate reports
            logging.info("Generating allocation reports...")
            summary_report = allocator.get_summary_report()
            logging.info(f"=== ALLOCATION SUMMARY ===")
            logging.info(f"Total Initial FTE: {summary_report['summary']['total_initial_fte']}")
            logging.info(f"Total Allocated FTE: {summary_report['summary']['total_allocated_fte']}")
            logging.info(f"Total Unutilized FTE: {summary_report['summary']['total_unutilized_fte']}")
            logging.info(f"Allocation Success Rate: {summary_report['summary']['allocation_success_rate']:.2%}")

            unmet_demand_df = allocator.get_unmet_demand_report()
            if not unmet_demand_df.empty:
                logging.warning(f"Found {len(unmet_demand_df)} allocation shortages")
                unmet_output_path = os.path.join(curpth, "unmet_demand_report.xlsx")
                unmet_demand_df.to_excel(unmet_output_path, index=False)
                logging.info(f"Saved unmet demand report to {unmet_output_path}")
            else:
                logging.info("No allocation shortages - all demand met!")

            unutilized_df = allocator.get_unutilized_report(consolidated_df)
            if not unutilized_df.empty:
                logging.info(f"Found {len(unutilized_df)} unutilized resource groups")
                unutilized_output_path = os.path.join(curpth, "unutilized_resources_report.xlsx")
                unutilized_df.to_excel(unutilized_output_path, index=False)
                logging.info(f"Saved unutilized resources report to {unutilized_output_path}")
            else:
                logging.info("All vendor resources utilized!")

            # Save processed data
            preprocessor = PreProcessing("forecast")
            mod_consolitated_df = preprocessor.preprocess_forecast_df(consolidated_df.copy())
            update_forecast_data(mod_consolitated_df, data_month, data_year, forecast_file_uploaded_by, forecast_filename)
            logging.info("Forecast data updated successfully.")
            update_summary_data(data_month, data_year)
            consolidated_df.to_excel(output_file)
            logging.info(f"Saved consolidated output to {output_file}")

            # SUCCESS: Complete execution tracking
            end_time = datetime.now()
            duration = (end_time - start_time).total_seconds()
            stats = {
                'records_processed': len(consolidated_df) if not consolidated_df.empty else 0,
                'allocation_success_rate': summary_report['summary'].get('allocation_success_rate', 0) if summary_report else 0
            }

            # Capture configuration snapshot for audit trail
            config_snapshot = build_config_snapshot(calculations)
            if config_snapshot.get('month_config'):
                update_status(execution_id, 'SUCCESS', config_snapshot=config_snapshot)
                logging.info(f"Captured config snapshot: {list(config_snapshot['month_config'].keys())}")

            # Cleanup old reports (retention policy - keep last 10 executions)
            try:
                db_manager.cleanup_old_reports(data_month, data_year, keep_last_n=10)
                logging.info("✓ Retention policy cleanup completed")
            except Exception as cleanup_error:
                logging.warning(f"Failed to cleanup old reports: {cleanup_error}")
                # Don't fail the execution if cleanup fails

            complete_execution(execution_id, success=True, stats=stats)
            logging.info(f"Processing completed successfully. Total time: {duration:.2f}s")

        else:
            # NO VALID DATAFRAMES: Mark as partial failure
            logging.error("No valid DataFrames generated. Check input files.")

            # Capture config snapshot if available
            try:
                config_snapshot = build_config_snapshot(calculations)
                if config_snapshot.get('month_config'):
                    update_status(execution_id, 'FAILED', config_snapshot=config_snapshot)
            except Exception as snapshot_error:
                logging.warning(f"Failed to capture config snapshot: {snapshot_error}")

            complete_execution(
                execution_id,
                success=False,
                error="No valid DataFrames generated. Check input files.",
                error_type='VALIDATION_ERROR'
            )

    except ValueError as e:
        # VALIDATION ERROR (Missing month config, missing columns, etc.)
        logging.error(f"Validation error during allocation: {e}", exc_info=True)

        # Capture config snapshot if available (may be empty if missing config caused the error)
        try:
            config_snapshot = build_config_snapshot(calculations)
            if config_snapshot.get('month_config'):
                update_status(execution_id, 'FAILED', config_snapshot=config_snapshot)
                logging.info(f"Captured partial config snapshot on error: {list(config_snapshot['month_config'].keys())}")
        except Exception as snapshot_error:
            logging.warning(f"Failed to capture config snapshot on ValueError: {snapshot_error}")

        complete_execution(
            execution_id,
            success=False,
            error=str(e),
            error_type='VALIDATION_ERROR',
            stack_trace=traceback.format_exc()
        )
        raise  # Re-raise to signal failure

    except Exception as e:
        # UNEXPECTED ERROR
        logging.error(f"Unexpected error during allocation: {e}", exc_info=True)

        # Capture config snapshot if available
        try:
            config_snapshot = build_config_snapshot(calculations)
            if config_snapshot.get('month_config'):
                update_status(execution_id, 'FAILED', config_snapshot=config_snapshot)
        except Exception as snapshot_error:
            logging.warning(f"Failed to capture config snapshot on Exception: {snapshot_error}")

        complete_execution(
            execution_id,
            success=False,
            error=str(e),
            error_type='UNEXPECTED_ERROR',
            stack_trace=traceback.format_exc()
        )
        raise  # Re-raise to signal failure

# Legacy function - no longer used with ResourceAllocator
# def process_row_level(row, month_headers, vendor_df):
#     for month in month_headers:
#         row[('FTE Avail', month)] = get_skills_split_count(row, month, vendor_df)
#         logging.debug(f"FTE Avail for {month}: {row[('FTE Avail', month)]}")
#     return row

def test_allocation_debug():
    """
    Simple test to verify allocation logic and see debug logs.
    Tests state mapping: matched states → specific buckets, unmatched → N/A bucket.
    """
    print("\n" + "="*80)
    print("TESTING RESOURCE ALLOCATOR - DEBUG MODE")
    print("="*80 + "\n")

    # Create minimal test vendor data
    # Note: "FL GA AR" - FL and GA match demand, AR doesn't (maps to N/A)
    #       "facets" - invalid state (maps to N/A)
    #       "AZ" - valid code but not in demand (maps to N/A)
    vendor_test_data = {
        'PrimaryPlatform': ['Amisys CROP', 'amisys', 'Facets', 'Amisys'],
        'State': ['FL GA AR', 'MI', 'facets', 'AZ'],
        'NewWorkType': ['FTC-Basic/Non MMP', 'ADJ-Basic/NON MMP', 'FTC MCARE', 'FTC-Basic/Non MMP'],
        'PartofProduction': ['Production', 'Ramp', 'Production', 'Production'],
        'BeelineTitle': ['Claims Analyst', 'Claims Analyst', 'Claims Analyst', 'Claims Analyst']
    }
    vendor_df = pd.DataFrame(vendor_test_data)

    # Create minimal test demand data
    # Demand states: FL, GA, MI, N/A
    demand_test_data = {
        ('Centene Capacity plan', 'Main LOB'): ['Amisys CROP', 'Amisys', 'Facets Global', 'Amisys'],
        ('Centene Capacity plan', 'State'): ['FL', 'MI', 'N/A', 'GA'],
        ('Centene Capacity plan', 'Case type'): ['FTC-Basic/Non MMP', 'ADJ-Basic/NON MMP', 'FTC MCARE', 'FTC-Basic/Non MMP']
    }
    demand_df = pd.DataFrame(demand_test_data)

    month_headers = ['March', 'April']

    print("\n--- TEST VENDOR DATA ---")
    print(vendor_df)
    print("\n--- TEST DEMAND DATA ---")
    print(demand_df)
    print("\n" + "="*80)
    print("EXPECTED STATE MAPPING:")
    print("  Vendor 'FL GA AR' → FL bucket, GA bucket, N/A bucket (AR unmapped)")
    print("  Vendor 'MI' → MI bucket (matched)")
    print("  Vendor 'facets' → N/A bucket (invalid)")
    print("  Vendor 'AZ' → N/A bucket (valid but not in demand)")
    print("="*80 + "\n")

    print("INITIALIZING RESOURCE ALLOCATOR (Watch for debug logs below)")
    print("="*80 + "\n")

    # Initialize allocator - this will trigger all debug logs
    allocator = ResourceAllocator(vendor_df, demand_df, month_headers)

    print("\n" + "="*80)
    print("TESTING ALLOCATION (First 5 requests will show detailed logs)")
    print("="*80 + "\n")

    # Test allocation - specific state match
    print("\n--- Test 1: Demand FL (should find FL bucket from 'FL GA AR' vendor) ---")
    allocated1, shortage1 = allocator.allocate('Amisys CROP', 'FL', 'March', 'FTC-Basic/Non MMP', 5)
    print(f"Result: Allocated={allocated1}, Shortage={shortage1}")

    # Test allocation - another specific state
    print("\n--- Test 2: Demand MI (should find MI bucket) ---")
    allocated2, shortage2 = allocator.allocate('Amisys', 'MI', 'March', 'ADJ-Basic/NON MMP', 3)
    print(f"Result: Allocated={allocated2}, Shortage={shortage2}")

    # Test allocation - N/A state (should access N/A buckets from AR, facets, AZ)
    print("\n--- Test 3: Demand N/A (should find N/A buckets from unmapped states) ---")
    allocated3, shortage3 = allocator.allocate('Facets Global', 'N/A', 'April', 'FTC MCARE', 2)
    print(f"Result: Allocated={allocated3}, Shortage={shortage3}")

    # Test allocation - GA state
    print("\n--- Test 4: Demand GA (should find GA bucket from 'FL GA AR' vendor) ---")
    allocated4, shortage4 = allocator.allocate('Amisys', 'GA', 'March', 'FTC-Basic/Non MMP', 2)
    print(f"Result: Allocated={allocated4}, Shortage={shortage4}")

    print("\n" + "="*80)
    print("TEST COMPLETE - Verify state mapping worked correctly:")
    print("  ✓ Matched states (FL, GA, MI) → specific buckets")
    print("  ✓ Unmatched states (AR, AZ, facets) → N/A bucket")
    print("  ✓ N/A demand can access N/A bucket resources")
    print("="*80 + "\n")

    return allocator

if __name__ == "__main__":
    # Uncomment to run full processing:
    # process_files('March', 2025, 'Makzoom Shah', 'NTT Forecast - v4_Capacity and HC_March_2025.xlsx')

    # Run debug test
    print("\n*** RUNNING DEBUG TEST MODE ***\n")
    test_allocation_debug()