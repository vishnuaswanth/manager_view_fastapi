"""
Bench Allocation Excel Export Module

Creates a 4-sheet Excel workbook with:
- Sheet 1: Summary
- Sheet 2: Modified Forecast Data (in download format)
- Sheet 3: Changes Detail
- Sheet 4: Vendor Assignments
"""

import os
from typing import Dict, List
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging

logger = logging.getLogger(__name__)


def create_changes_workbook(
    changes: List[Dict],
    summary: Dict,
    modified_forecast_rows: pd.DataFrame,
    vendor_assignments: List[Dict],
    output_dir: str = "/tmp"
) -> str:
    """
    Create Excel workbook with bench allocation results.

    Args:
        changes: List of change dicts with before/after values
        summary: Summary statistics dict
        modified_forecast_rows: DataFrame with modified forecast data in download format
        vendor_assignments: List of vendor assignment dicts
        output_dir: Directory to save the Excel file

    Returns:
        str: Path to the created Excel file
    """
    try:
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        month = summary.get('month', 'Unknown')
        year = summary.get('year', 'Unknown')
        filename = f"bench_allocation_{month}_{year}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Create sheets
        _create_summary_sheet(wb, summary)
        _create_modified_forecast_sheet(wb, modified_forecast_rows)
        _create_changes_detail_sheet(wb, changes)
        _create_vendor_assignments_sheet(wb, vendor_assignments)

        # Save workbook
        wb.save(filepath)
        logger.info(f"Created Excel workbook: {filepath}")

        return filepath

    except Exception as e:
        logger.error(f"Error creating Excel workbook: {e}", exc_info=True)
        raise


def _create_summary_sheet(wb: Workbook, summary: Dict):
    """
    Create Sheet 1: Summary with allocation statistics.
    """
    ws = wb.create_sheet("Summary", 0)

    # Header styling
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")

    # Title
    ws['A1'] = "Bench Allocation Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')

    # Basic info
    row = 3
    ws[f'A{row}'] = "Month:"
    ws[f'B{row}'] = summary.get('month', 'N/A')
    ws[f'A{row}'].font = header_font

    row += 1
    ws[f'A{row}'] = "Year:"
    ws[f'B{row}'] = summary.get('year', 'N/A')
    ws[f'A{row}'].font = header_font

    row += 1
    ws[f'A{row}'] = "Generation Timestamp:"
    ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws[f'A{row}'].font = header_font

    row += 1
    ws[f'A{row}'] = "Allocation Strategy:"
    ws[f'B{row}'] = "Fill Gaps First, Then Proportional (Whole FTEs)"
    ws[f'A{row}'].font = header_font

    # Allocation Statistics
    row += 2
    ws[f'A{row}'] = "Allocation Statistics"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')

    row += 1
    ws[f'A{row}'] = "Total Bench Allocated:"
    ws[f'B{row}'] = summary.get('total_bench_allocated', 0)
    ws[f'A{row}'].font = header_font

    row += 1
    ws[f'A{row}'] = "Gaps Filled:"
    ws[f'B{row}'] = summary.get('gaps_filled', 0)
    ws[f'A{row}'].font = header_font

    row += 1
    ws[f'A{row}'] = "Excess Distributed:"
    ws[f'B{row}'] = summary.get('excess_distributed', 0)
    ws[f'A{row}'].font = header_font

    row += 1
    ws[f'A{row}'] = "Rows Modified:"
    ws[f'B{row}'] = summary.get('rows_modified', 0)
    ws[f'A{row}'].font = header_font

    # Validation Status
    row += 2
    ws[f'A{row}'] = "Validation Status"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')

    row += 1
    ws[f'A{row}'] = "Allocation Valid:"
    validation_result = summary.get('validation', {})
    is_valid = validation_result.get('valid', False)
    ws[f'B{row}'] = "Yes" if is_valid else "No"
    ws[f'A{row}'].font = header_font
    if not is_valid:
        ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if 'error' in validation_result:
        row += 1
        ws[f'A{row}'] = "Validation Error:"
        ws[f'B{row}'] = validation_result.get('error', '')
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Warnings
    warnings = summary.get('warnings', [])
    if warnings:
        row += 2
        ws[f'A{row}'] = "Warnings"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')

        for warning in warnings:
            row += 1
            ws[f'A{row}'] = warning
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Auto-size columns
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50


def _create_modified_forecast_sheet(wb: Workbook, modified_forecast_rows: pd.DataFrame):
    """
    Create Sheet 2: Modified Forecast Data in download format.
    Only includes rows that were modified during allocation.
    """
    ws = wb.create_sheet("Modified Forecast Data")

    if modified_forecast_rows.empty:
        ws['A1'] = "No forecast rows were modified"
        ws['A1'].font = Font(italic=True)
        return

    # Define column headers matching ForecastModel download format
    column_headers = [
        'Centene Capacity Plan Main LOB',
        'Centene Capacity Plan State',
        'Centene Capacity Plan Case Type',
        'Centene Capacity Plan Call Type ID',
        'Centene Capacity Plan Target CPH',
        'Client Forecast Month1',
        'Client Forecast Month2',
        'Client Forecast Month3',
        'Client Forecast Month4',
        'Client Forecast Month5',
        'Client Forecast Month6',
        'FTE Required Month1',
        'FTE Required Month2',
        'FTE Required Month3',
        'FTE Required Month4',
        'FTE Required Month5',
        'FTE Required Month6',
        'FTE Avail Month1',
        'FTE Avail Month2',
        'FTE Avail Month3',
        'FTE Avail Month4',
        'FTE Avail Month5',
        'FTE Avail Month6',
        'Capacity Month1',
        'Capacity Month2',
        'Capacity Month3',
        'Capacity Month4',
        'Capacity Month5',
        'Capacity Month6',
        'Month',
        'Year',
        'UploadedFile',
        'UpdatedBy',
        'UpdatedDateTime'
    ]

    # Write headers
    for col_idx, header in enumerate(column_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_idx, row in enumerate(dataframe_to_rows(modified_forecast_rows, index=False, header=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Highlight modified FTE_Avail and Capacity cells (columns 18-29)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    for row_idx in range(2, ws.max_row + 1):
        # Highlight FTE_Avail columns (18-23)
        for col_idx in range(18, 24):
            ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

        # Highlight Capacity columns (24-29)
        for col_idx in range(24, 30):
            ws.cell(row=row_idx, column=col_idx).fill = orange_fill

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Auto-size columns
    for col_idx in range(1, len(column_headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20


def _create_changes_detail_sheet(wb: Workbook, changes: List[Dict]):
    """
    Create Sheet 3: Changes Detail showing before/after values.
    """
    ws = wb.create_sheet("Changes Detail")

    if not changes:
        ws['A1'] = "No changes were made"
        ws['A1'].font = Font(italic=True)
        return

    # Create DataFrame from changes
    changes_df = pd.DataFrame(changes)

    # Define column order
    columns = [
        'main_lob',
        'state',
        'case_type',
        'month',
        'fte_required',
        'fte_avail_before',
        'fte_avail_after',
        'fte_change',
        'allocation_type',
        'vendors_allocated',
        'capacity_before',
        'capacity_after',
        'capacity_change',
        'capacity_pct_change'
    ]

    # Rename columns for display
    display_columns = {
        'main_lob': 'Main LOB',
        'state': 'State',
        'case_type': 'Case Type',
        'month': 'Month',
        'fte_required': 'FTE Required',
        'fte_avail_before': 'FTE Avail (Before)',
        'fte_avail_after': 'FTE Avail (After)',
        'fte_change': 'Change (+)',
        'allocation_type': 'Allocation Type',
        'vendors_allocated': 'Vendors Allocated',
        'capacity_before': 'Capacity (Before)',
        'capacity_after': 'Capacity (After)',
        'capacity_change': 'Capacity Change',
        'capacity_pct_change': 'Capacity % Change'
    }

    # Reorder and rename
    changes_df = changes_df[columns]
    changes_df = changes_df.rename(columns=display_columns)

    # Write headers
    for col_idx, col_name in enumerate(changes_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, row in enumerate(dataframe_to_rows(changes_df, index=False, header=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Apply row highlighting based on allocation type
    gap_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    excess_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Light blue

    allocation_type_col = list(display_columns.values()).index('Allocation Type') + 1

    for row_idx in range(2, ws.max_row + 1):
        allocation_type = ws.cell(row=row_idx, column=allocation_type_col).value
        if allocation_type == "Gap Fill":
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = gap_fill
        elif allocation_type == "Excess":
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = excess_fill

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Auto-size columns
    for col_idx in range(1, len(display_columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18


def _create_vendor_assignments_sheet(wb: Workbook, vendor_assignments: List[Dict]):
    """
    Create Sheet 4: Vendor Assignments showing which vendors were allocated where.
    """
    ws = wb.create_sheet("Vendor Assignments")

    if not vendor_assignments:
        ws['A1'] = "No vendors were assigned"
        ws['A1'].font = Font(italic=True)
        return

    # Create DataFrame from vendor assignments
    assignments_df = pd.DataFrame(vendor_assignments)

    # Define column order
    columns = [
        'vendor_name',
        'vendor_cn',
        'vendor_skills',
        'vendor_states',
        'allocated_to_lob',
        'allocated_to_state',
        'allocated_to_worktype',
        'allocation_month',
        'allocation_type'
    ]

    # Rename columns for display
    display_columns = {
        'vendor_name': 'Vendor Name',
        'vendor_cn': 'Vendor CN',
        'vendor_skills': 'Vendor Skills',
        'vendor_states': 'Vendor States',
        'allocated_to_lob': 'Allocated To LOB',
        'allocated_to_state': 'Allocated To State',
        'allocated_to_worktype': 'Allocated To Worktype',
        'allocation_month': 'Allocation Month',
        'allocation_type': 'Allocation Type'
    }

    # Reorder and rename
    assignments_df = assignments_df[columns]
    assignments_df = assignments_df.rename(columns=display_columns)

    # Write headers
    for col_idx, col_name in enumerate(assignments_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, row in enumerate(dataframe_to_rows(assignments_df, index=False, header=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Auto-size columns
    for col_idx in range(1, len(display_columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
