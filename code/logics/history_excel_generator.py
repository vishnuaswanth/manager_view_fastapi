"""
History Excel Export Generator.

Generates Excel files for history log downloads with pivot table format
showing before/after values for forecast changes.
"""

import logging
import pandas as pd
from io import BytesIO
from typing import Dict, List, Optional, Union, Any, Tuple
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)

# Core field names (always in this order under each month)
CORE_FIELDS = ["Client Forecast", "FTE Required", "FTE Available", "Capacity"]


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def _parse_month_label(month_label: str) -> tuple:
    """
    Parse month label for chronological sorting.

    Args:
        month_label: Month label like "Jun-25" or "Jun-2025"

    Returns:
        Tuple of (year, month_num) for sorting

    Example:
        >>> _parse_month_label("Jun-25")
        (2025, 6)
        >>> _parse_month_label("Dec-24")
        (2024, 12)
    """
    import re
    from datetime import datetime

    # Try to extract month and year
    match = re.match(r'([A-Za-z]{3})-(\d{2,4})', month_label)
    if match:
        month_str, year_str = match.groups()

        # Parse month name to number
        try:
            month_num = datetime.strptime(month_str, '%b').month
        except ValueError:
            # If parsing fails, return as-is
            return (9999, 99)

        # Parse year (handle 2-digit or 4-digit)
        year = int(year_str)
        if year < 100:
            year += 2000

        return (year, month_num)

    # Fallback: return as-is for alphabetical sort
    return (9999, 99)  # Sort unknown formats to end


# ============================================================================
# TYPE-SAFE DATA STRUCTURES
# ============================================================================

@dataclass
class HistoryChangeRecord:
    """
    Type-safe structure for a single history change record.

    Represents one field-level change (e.g., "Jun-25.fte_avail" changed from 20 to 25).

    Attributes:
        main_lob: Main Line of Business identifier
        state: State code (2-letter)
        case_type: Case type description
        case_id: Business case identifier (Centene_Capacity_Plan_Call_Type_ID)
        field_name: Field path in DOT notation (e.g., "Jun-25.fte_avail" or "target_cph")
        old_value: Previous value before change (can be None for new records)
        new_value: Current value after change (can be None for deleted records)
        delta: Change amount (new_value - old_value), can be None
        month_label: Month label if field is month-specific (e.g., "Jun-25"), None otherwise

    Example:
        >>> change = HistoryChangeRecord(
        ...     main_lob="Amisys Medicaid DOMESTIC",
        ...     state="TX",
        ...     case_type="Claims Processing",
        ...     case_id="CL-001",
        ...     field_name="Jun-25.fte_avail",
        ...     old_value="20",
        ...     new_value="25",
        ...     delta=5.0,
        ...     month_label="Jun-25"
        ... )
    """
    main_lob: str
    state: str
    case_type: str
    case_id: str
    field_name: str
    old_value: Optional[Union[str, int, float]]
    new_value: Optional[Union[str, int, float]]
    delta: Optional[Union[int, float]]
    month_label: Optional[str]

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'HistoryChangeRecord':
        """
        Create HistoryChangeRecord from dict.

        Args:
            data: Dict with change record data

        Returns:
            HistoryChangeRecord instance

        Raises:
            KeyError: If required keys are missing
            ValueError: If data types are invalid
        """
        required_keys = ['main_lob', 'state', 'case_type', 'case_id', 'field_name']
        missing_keys = [key for key in required_keys if key not in data]
        if missing_keys:
            raise KeyError(f"Missing required keys in change record: {missing_keys}")

        return cls(
            main_lob=str(data['main_lob']),
            state=str(data['state']),
            case_type=str(data['case_type']),
            case_id=str(data['case_id']),
            field_name=str(data['field_name']),
            old_value=data.get('old_value'),
            new_value=data.get('new_value'),
            delta=data.get('delta'),
            month_label=data.get('month_label')
        )

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dict for backward compatibility."""
        return {
            'main_lob': self.main_lob,
            'state': self.state,
            'case_type': self.case_type,
            'case_id': self.case_id,
            'field_name': self.field_name,
            'old_value': self.old_value,
            'new_value': self.new_value,
            'delta': self.delta,
            'month_label': self.month_label
        }


@dataclass
class SummaryTotals:
    """
    Type-safe structure for summary totals by month.

    Attributes:
        old: Old total value
        new: New total value
    """
    old: Union[int, float]
    new: Union[int, float]


@dataclass
class MonthSummary:
    """
    Type-safe structure for month-level summary data.

    Attributes:
        total_forecast: Optional forecast totals (old/new)
        total_fte_required: Optional FTE required totals (old/new)
        total_fte_available: Optional FTE available totals (old/new)
        total_capacity: Optional capacity totals (old/new)
    """
    total_forecast: Optional[SummaryTotals] = None
    total_fte_required: Optional[SummaryTotals] = None
    total_fte_available: Optional[SummaryTotals] = None
    total_capacity: Optional[SummaryTotals] = None


@dataclass
class HistorySummaryData:
    """
    Type-safe structure for history log summary data.

    Attributes:
        report_month: Report month name (e.g., "March")
        report_year: Report year (e.g., 2025)
        months: List of month labels included (e.g., ["Jun-25", "Jul-25"])
        totals: Dict mapping month labels to MonthSummary objects
    """
    report_month: str
    report_year: int
    months: List[str]
    totals: Dict[str, MonthSummary]

    @classmethod
    def from_dict(cls, data: Optional[Dict[str, Any]]) -> Optional['HistorySummaryData']:
        """
        Create HistorySummaryData from dict.

        Args:
            data: Optional dict with summary data

        Returns:
            HistorySummaryData instance or None if data is None/empty
        """
        if not data:
            return None

        # Parse totals
        totals_dict = {}
        if 'totals' in data:
            for month_label, month_data in data['totals'].items():
                month_summary = MonthSummary()

                # Parse each metric
                if 'total_forecast' in month_data:
                    month_summary.total_forecast = SummaryTotals(
                        old=month_data['total_forecast'].get('old', 0),
                        new=month_data['total_forecast'].get('new', 0)
                    )
                if 'total_fte_required' in month_data:
                    month_summary.total_fte_required = SummaryTotals(
                        old=month_data['total_fte_required'].get('old', 0),
                        new=month_data['total_fte_required'].get('new', 0)
                    )
                if 'total_fte_available' in month_data:
                    month_summary.total_fte_available = SummaryTotals(
                        old=month_data['total_fte_available'].get('old', 0),
                        new=month_data['total_fte_available'].get('new', 0)
                    )
                if 'total_capacity' in month_data:
                    month_summary.total_capacity = SummaryTotals(
                        old=month_data['total_capacity'].get('old', 0),
                        new=month_data['total_capacity'].get('new', 0)
                    )

                totals_dict[month_label] = month_summary

        return cls(
            report_month=data.get('report_month', ''),
            report_year=int(data.get('report_year', 0)),
            months=data.get('months', []),
            totals=totals_dict
        )


@dataclass
class HistoryLogData:
    """
    Type-safe structure for history log metadata.

    Attributes:
        id: History log UUID
        change_type: Type of change (e.g., "Bench Allocation", "CPH Update")
        month: Report month name (e.g., "March")
        year: Report year (e.g., 2025)
        timestamp: ISO timestamp string
        user: User who made the change
        description: Optional user notes/description
        records_modified: Count of modified records
        summary_data: Optional aggregated summary data
    """
    id: str
    change_type: str
    month: str
    year: int
    timestamp: str
    user: str
    description: Optional[str]
    records_modified: int
    summary_data: Optional[HistorySummaryData]

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'HistoryLogData':
        """
        Create HistoryLogData from dict.

        Args:
            data: Dict with history log data

        Returns:
            HistoryLogData instance

        Raises:
            KeyError: If required keys are missing
            ValueError: If data types are invalid
        """
        required_keys = ['id', 'change_type', 'month', 'year', 'timestamp', 'user', 'records_modified']
        missing_keys = [key for key in required_keys if key not in data]
        if missing_keys:
            raise KeyError(f"Missing required keys in history log data: {missing_keys}")

        return cls(
            id=str(data['id']),
            change_type=str(data['change_type']),
            month=str(data['month']),
            year=int(data['year']),
            timestamp=str(data['timestamp']),
            user=str(data['user']),
            description=str(data['description']) if data.get('description') else None,
            records_modified=int(data['records_modified']),
            summary_data=HistorySummaryData.from_dict(data.get('summary_data'))
        )

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dict for backward compatibility."""
        result = {
            'id': self.id,
            'change_type': self.change_type,
            'month': self.month,
            'year': self.year,
            'timestamp': self.timestamp,
            'user': self.user,
            'description': self.description,
            'records_modified': self.records_modified
        }

        if self.summary_data:
            result['summary_data'] = {
                'report_month': self.summary_data.report_month,
                'report_year': self.summary_data.report_year,
                'months': self.summary_data.months,
                'totals': {
                    month_label: {
                        'total_forecast': {'old': month_summary.total_forecast.old, 'new': month_summary.total_forecast.new} if month_summary.total_forecast else None,
                        'total_fte_required': {'old': month_summary.total_fte_required.old, 'new': month_summary.total_fte_required.new} if month_summary.total_fte_required else None,
                        'total_fte_available': {'old': month_summary.total_fte_available.old, 'new': month_summary.total_fte_available.new} if month_summary.total_fte_available else None,
                        'total_capacity': {'old': month_summary.total_capacity.old, 'new': month_summary.total_capacity.new} if month_summary.total_capacity else None,
                    }
                    for month_label, month_summary in self.summary_data.totals.items()
                }
            }

        return result


def generate_history_excel(
    history_log_data: HistoryLogData,
    changes: List[HistoryChangeRecord]
) -> BytesIO:
    """
    Generate Excel file for history log download.

    IMPORTANT: This function expects type-safe dataclass objects, not raw dicts.
    Use HistoryLogData.from_dict() and HistoryChangeRecord.from_dict() to convert
    if you have dict data.

    Args:
        history_log_data: History log metadata (HistoryLogData instance)
        changes: List of field-level changes (List[HistoryChangeRecord])

    Returns:
        BytesIO: Excel file buffer

    Raises:
        ValueError: If data is invalid or changes list is empty
        TypeError: If parameters are not the expected types

    Example:
        # Convert dict data to type-safe objects
        history_log = HistoryLogData.from_dict(history_data)
        typed_changes = [HistoryChangeRecord.from_dict(c) for c in changes_data]

        # Generate Excel
        excel_buffer = generate_history_excel(history_log, typed_changes)
    """
    # Validate types
    if not isinstance(history_log_data, HistoryLogData):
        raise TypeError(
            f"history_log_data must be HistoryLogData instance, got {type(history_log_data)}. "
            f"Use HistoryLogData.from_dict() to convert from dict."
        )

    if not isinstance(changes, list):
        raise TypeError(f"changes must be a list, got {type(changes)}")

    if not changes:
        raise ValueError("No changes provided for Excel generation")

    # Validate all changes are HistoryChangeRecord instances
    for i, change in enumerate(changes):
        if not isinstance(change, HistoryChangeRecord):
            raise TypeError(
                f"Change at index {i} must be HistoryChangeRecord instance, got {type(change)}. "
                f"Use HistoryChangeRecord.from_dict() to convert from dict."
            )

    typed_changes = changes  # Already validated as List[HistoryChangeRecord]

    # Step 1: Transform changes to pivot table structure WITH metadata
    pivot_data, month_labels, static_columns = _prepare_pivot_data(typed_changes)

    # Step 2: Create Excel workbook with openpyxl
    excel_buffer = BytesIO()

    # Step 3: Build correct column order to match header structure
    # Order: [Static columns] [Month1 fields] [Month2 fields] ...
    column_order = list(static_columns)  # Start with static columns

    # Add month columns in order (chronological, already sorted)
    for month_label in month_labels:
        for field in CORE_FIELDS:
            column_order.append(f"{month_label} {field}")

    logger.debug(f"Expected column order for DataFrame: {column_order}")

    # Validate that all expected columns exist in pivot_data
    if pivot_data:
        actual_columns = set(pivot_data[0].keys())
        expected_columns = set(column_order)

        missing_columns = expected_columns - actual_columns
        extra_columns = actual_columns - expected_columns

        if missing_columns:
            logger.warning(f"Missing columns in pivot_data: {missing_columns}")
        if extra_columns:
            logger.warning(f"Extra columns in pivot_data (will be ignored): {extra_columns}")

    # Step 4: Create main data sheet with explicit column order
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        # Write pivot data with explicit column order
        df_pivot = pd.DataFrame(pivot_data, columns=column_order)
        # Write WITHOUT headers - we'll create custom multi-level headers manually
        # Start data at row 3 (rows 1-2 for headers)
        df_pivot.to_excel(writer, sheet_name='Changes', index=False, header=False, startrow=2)

        # Write summary sheet
        summary_data = _prepare_summary_sheet(history_log_data)
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False, header=False)

    # Step 5: Apply multi-level headers and formatting
    excel_buffer.seek(0)
    _apply_multilevel_headers_and_formatting(
        excel_buffer,
        history_log_data,
        month_labels,
        static_columns
    )

    excel_buffer.seek(0)
    logger.info(f"Generated Excel for history log {history_log_data.id}")
    return excel_buffer


def _prepare_pivot_data(changes: List[HistoryChangeRecord]) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
    """
    Transform changes list to pivot table structure.

    Groups changes by forecast record (main_lob, state, case_type, case_id)
    and creates columns for each month+field combination.

    Args:
        changes: List of HistoryChangeRecord objects

    Returns:
        Tuple of:
        - pivot_rows: List of data row dicts
        - month_labels: Ordered list of unique month labels (e.g., ["Jun-25", "Jul-25"])
        - static_columns: List of static column names (e.g., ["Main LOB", "State", ...])

    Raises:
        ValueError: If changes contain invalid field names
    """
    # Group changes by forecast record
    grouped_changes: Dict[tuple, Dict[str, Any]] = {}

    # Track month labels encountered
    month_labels_set = set()

    for change in changes:
        # Create composite key for grouping
        key = (
            change.main_lob,
            change.state,
            change.case_type,
            change.case_id
        )

        if key not in grouped_changes:
            grouped_changes[key] = {
                'main_lob': change.main_lob,
                'state': change.state,
                'case_type': change.case_type,
                'case_id': change.case_id,
                'fields': {}
            }

        # Parse field path
        field_name = change.field_name

        # Validate field_name
        if not field_name:
            logger.warning(f"Empty field_name for change: {change.case_id}")
            continue

        # Determine display name
        if '.' in field_name:
            # Month-specific field: "Jun-25.fte_avail"
            parts = field_name.split('.', 1)
            if len(parts) != 2:
                logger.warning(f"Invalid field_name format: {field_name}")
                continue

            month_label, metric = parts
            month_labels_set.add(month_label)  # Track month label
            metric_display = _get_metric_display_name(metric)
            column_name = f"{month_label} {metric_display}"
        else:
            # Month-agnostic field: "target_cph"
            column_name = _get_metric_display_name(field_name)

        # Format value (show old in brackets if changed)
        old_val = change.old_value
        new_val = change.new_value

        if old_val is not None and new_val is not None and str(old_val) != str(new_val):
            display_value = f"{new_val} ({old_val})"
        else:
            display_value = new_val if new_val is not None else old_val

        grouped_changes[key]['fields'][column_name] = display_value

    # Convert to list of row dicts
    pivot_rows = []
    for (main_lob, state, case_type, case_id), record in grouped_changes.items():
        row = {
            'Main LOB': main_lob,
            'State': state,
            'Case Type': case_type,
            'Case ID': case_id
        }
        row.update(record['fields'])
        pivot_rows.append(row)

    # Sort month labels chronologically
    month_labels = sorted(month_labels_set, key=_parse_month_label)

    # Determine static columns
    static_columns = ["Main LOB", "State", "Case Type", "Case ID"]

    # Check if Target CPH exists in any record
    has_target_cph = any("Target CPH" in row for row in pivot_rows)
    if has_target_cph:
        static_columns.append("Target CPH")

    logger.debug(f"Prepared {len(pivot_rows)} pivot rows from {len(changes)} changes")
    logger.debug(f"Month labels: {month_labels}")
    logger.debug(f"Static columns: {static_columns}")

    return (pivot_rows, month_labels, static_columns)


def _get_metric_display_name(metric: str) -> str:
    """
    Convert API metric name to display name.

    Args:
        metric: API metric name (e.g., "fte_avail", "target_cph")

    Returns:
        Display-friendly metric name (e.g., "FTE Available", "Target CPH")

    Example:
        >>> _get_metric_display_name("fte_avail")
        "FTE Available"
        >>> _get_metric_display_name("target_cph")
        "Target CPH"
        >>> _get_metric_display_name("custom_field")
        "Custom Field"
    """
    metric_map = {
        'forecast': 'Client Forecast',
        'fte_req': 'FTE Required',
        'fte_avail': 'FTE Available',
        'capacity': 'Capacity',
        'target_cph': 'Target CPH'
    }
    return metric_map.get(metric, metric.replace('_', ' ').title())


def _create_multilevel_headers(
    ws,
    month_labels: List[str],
    static_columns: List[str],
    core_fields: List[str]
) -> None:
    """
    Create two-level headers with merged cells.

    Structure:
    Row 1: [Static cols (merged)] [Month1 (merged 4 cols)] [Month2 (merged 4 cols)] ...
    Row 2: [Static cols (merged)] [Field1] [Field2] [Field3] [Field4] [Field1] ...

    NOTE: This function assumes rows 1-2 are empty (pandas data starts at row 3).
    It directly populates rows 1-2 without inserting any rows.

    KNOWN BEHAVIOR: Excel for Mac may show a repair dialog when opening the file.
    This is due to openpyxl's merged cell handling removing non-anchor cells from XML,
    which Excel for Mac's strict validation flags as invalid. The file is functionally
    correct and will open successfully after clicking "Yes" on the repair dialog.
    All data and formatting will be preserved correctly.

    Args:
        ws: openpyxl Worksheet (with empty rows 1-2, data starting at row 3)
        month_labels: List of month labels (e.g., ["Jun-25", "Jul-25"])
        static_columns: List of static column names
        core_fields: List of field names under each month
            (e.g., ["Client Forecast", "FTE Required", "FTE Available", "Capacity"])
    """
    # NOTE: Rows 1-2 are empty, data starts at row 3
    # We populate rows 1-2 directly without inserting

    # Current column index
    col_idx = 1
    total_merges = 0

    # 1. Create static column headers (merged vertically)
    for col_name in static_columns:
        # Write value ONLY to row 1 (anchor cell)
        ws.cell(row=1, column=col_idx, value=col_name)

        # DON'T write to row 2 - leave empty for clean merge

        # Merge
        end_col_letter = get_column_letter(col_idx)
        ws.merge_cells(f'{end_col_letter}1:{end_col_letter}2')
        total_merges += 1
        col_idx += 1

    # 2. Create month headers with field subheaders
    for month_label in month_labels:
        start_col = col_idx
        end_col = col_idx + len(core_fields) - 1

        # Write month value ONLY to anchor cell in row 1
        ws.cell(row=1, column=start_col, value=month_label)

        # Merge across columns in row 1
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        ws.merge_cells(f'{start_letter}1:{end_letter}1')
        total_merges += 1

        # Write field names to row 2 (no merge)
        for field_name in core_fields:
            ws.cell(row=2, column=col_idx, value=field_name)
            col_idx += 1

    logger.debug(f"Created multi-level headers for {len(month_labels)} months with {total_merges} merged ranges")


def _prepare_summary_sheet(history_log_data: HistoryLogData) -> List[Dict[str, str]]:
    """
    Prepare summary sheet data.

    Args:
        history_log_data: HistoryLogData instance with metadata

    Returns:
        List of {label: value} dicts for vertical layout
    """
    summary_rows = [
        {'label': 'History Log ID', 'value': history_log_data.id},
        {'label': 'Change Type', 'value': history_log_data.change_type},
        {'label': 'Report Month', 'value': f"{history_log_data.month} {history_log_data.year}"},
        {'label': 'Timestamp', 'value': history_log_data.timestamp},
        {'label': 'User', 'value': history_log_data.user},
        {'label': 'Description', 'value': history_log_data.description or 'N/A'},
        {'label': 'Records Modified', 'value': str(history_log_data.records_modified)},
    ]

    # Add summary data if present
    if history_log_data.summary_data:
        summary_rows.append({'label': '', 'value': ''})  # Blank row
        summary_rows.append({'label': 'SUMMARY TOTALS', 'value': ''})

        summary_data = history_log_data.summary_data

        # Handle both dict and HistorySummaryData object
        if isinstance(summary_data, dict):
            totals = summary_data.get('totals', {})
        else:
            totals = summary_data.totals if hasattr(summary_data, 'totals') else {}

        if totals:
            for month_label, month_summary in totals.items():
                # Handle both dict and object for month_summary
                if isinstance(month_summary, dict):
                    # Dict format from database
                    fte_avail = month_summary.get('total_fte_available')
                    forecast = month_summary.get('total_forecast')
                    capacity = month_summary.get('total_capacity')
                else:
                    # Object format from typed conversion
                    fte_avail = month_summary.total_fte_available if hasattr(month_summary, 'total_fte_available') else None
                    forecast = month_summary.total_forecast if hasattr(month_summary, 'total_forecast') else None
                    capacity = month_summary.total_capacity if hasattr(month_summary, 'total_capacity') else None

                # Add FTE Available totals if present
                if fte_avail:
                    old_val = fte_avail.get('old') if isinstance(fte_avail, dict) else fte_avail.old
                    new_val = fte_avail.get('new') if isinstance(fte_avail, dict) else fte_avail.new
                    summary_rows.append({
                        'label': f"{month_label} Total FTE Available (Old)",
                        'value': str(old_val)
                    })
                    summary_rows.append({
                        'label': f"{month_label} Total FTE Available (New)",
                        'value': str(new_val)
                    })

                # Add other metrics if present
                if forecast:
                    old_val = forecast.get('old') if isinstance(forecast, dict) else forecast.old
                    new_val = forecast.get('new') if isinstance(forecast, dict) else forecast.new
                    summary_rows.append({
                        'label': f"{month_label} Total Forecast (Old)",
                        'value': str(old_val)
                    })
                    summary_rows.append({
                        'label': f"{month_label} Total Forecast (New)",
                        'value': str(new_val)
                    })

                if capacity:
                    old_val = capacity.get('old') if isinstance(capacity, dict) else capacity.old
                    new_val = capacity.get('new') if isinstance(capacity, dict) else capacity.new
                    summary_rows.append({
                        'label': f"{month_label} Total Capacity (Old)",
                        'value': str(old_val)
                    })
                    summary_rows.append({
                        'label': f"{month_label} Total Capacity (New)",
                        'value': str(new_val)
                    })

    logger.debug(f"Prepared {len(summary_rows)} summary rows")
    return summary_rows


def _apply_multilevel_headers_and_formatting(
    excel_buffer: BytesIO,
    history_log_data: HistoryLogData,
    month_labels: List[str],
    static_columns: List[str]
) -> None:
    """
    Apply multi-level headers and formatting to Excel workbook.

    Args:
        excel_buffer: BytesIO with Excel data
        history_log_data: HistoryLogData instance for metadata
        month_labels: List of month labels for multi-level headers
        static_columns: List of static column names

    Raises:
        Exception: If workbook loading or formatting fails
    """
    try:
        wb = load_workbook(excel_buffer)
    except Exception as e:
        logger.error(f"Failed to load workbook for formatting: {e}", exc_info=True)
        raise

    # Format Changes sheet
    if 'Changes' in wb.sheetnames:
        ws_changes = wb['Changes']

        # STEP 1: Create multi-level headers (inserts row 1, shifts data down)
        _create_multilevel_headers(
            ws_changes,
            month_labels,
            static_columns,
            CORE_FIELDS
        )

        # STEP 2: Define styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        month_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        field_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # STEP 3: Style row 1 (month headers + static columns)
        for cell in ws_changes[1]:
            if not isinstance(cell, MergedCell):
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
                cell.fill = month_header_fill  # Dark blue for all row 1 headers

        # STEP 4: Style row 2 (field headers, including merged static cells)
        for cell in ws_changes[2]:
            # Skip if part of merged cell from row 1
            if not isinstance(cell, MergedCell):
                cell.font = header_font
                cell.fill = field_header_fill  # Lighter blue for fields
                cell.alignment = header_alignment
                cell.border = border

        # STEP 5: Apply borders and alignment to data rows (starts at row 3 now)
        for row in ws_changes.iter_rows(min_row=3, max_row=ws_changes.max_row, max_col=ws_changes.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top')

        # STEP 6: Auto-size columns
        for column in ws_changes.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws_changes.column_dimensions[column_letter].width = adjusted_width

    # Format Summary sheet
    if 'Summary' in wb.sheetnames:
        ws_summary = wb['Summary']

        # Bold first column (labels)
        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row):
            row[0].font = Font(bold=True)

        # Auto-size
        for column in ws_summary.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws_summary.column_dimensions[column_letter].width = max_length + 2

    # Save changes
    wb.save(excel_buffer)
    logger.info("Applied multi-level headers and formatting successfully")
