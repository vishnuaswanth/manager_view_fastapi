"""
Final test: Create an Excel file exactly as the API would, and verify it's valid.
"""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    HistoryLogData,
    HistoryChangeRecord,
    generate_history_excel
)
from openpyxl import load_workbook
import tempfile


def save_and_reopen_excel(excel_buffer):
    """Save Excel to temp file and reopen to test validity."""
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        tmp.write(excel_buffer.getvalue())
        tmp_path = tmp.name

    print(f"  Saved to: {tmp_path}")

    # Try to open with openpyxl
    try:
        wb = load_workbook(tmp_path)
        print(f"  ✓ openpyxl can open file")
        return wb, tmp_path
    except Exception as e:
        print(f"  ✗ openpyxl FAILED to open: {e}")
        raise


def test_realistic_bench_allocation():
    """Test with realistic bench allocation data (Option 1)."""
    print("\n" + "=" * 70)
    print("Test: Realistic Bench Allocation (Option 1 - All Fields)")
    print("=" * 70)

    history_log = HistoryLogData(
        id="ba-12345",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T14:30:00Z",
        user="system",
        description="Automated bench allocation for March 2025",
        records_modified=1,
        summary_data={
            'report_month': 'March',
            'report_year': 2025,
            'months': ['Jun-25'],
            'totals': {
                'Jun-25': {
                    'total_fte_available': {'old': 100, 'new': 105}
                }
            }
        }
    )

    # Option 1: Track ALL 4 fields even if only 2 changed
    # This is what bench_allocation_transformer creates
    changes = [
        # Forecast - unchanged (Option 1 tracks it anyway)
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.forecast",
            old_value=1000,
            new_value=1000,
            delta=0,
            month_label="Jun-25"
        ),
        # FTE Required - unchanged (Option 1 tracks it anyway)
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.fte_req",
            old_value=20,
            new_value=20,
            delta=0,
            month_label="Jun-25"
        ),
        # FTE Available - CHANGED (actual bench allocation)
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,
            delta=5,
            month_label="Jun-25"
        ),
        # Capacity - CHANGED (recalculated)
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims Processing",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1125,
            delta=125,
            month_label="Jun-25"
        )
    ]

    print("\nGenerating Excel...")
    excel_buffer = generate_history_excel(history_log, changes)

    print("  ✓ Excel generated")

    print("\nSaving and reopening...")
    wb, tmp_path = save_and_reopen_excel(excel_buffer)

    print("\nValidating structure...")
    ws = wb['Changes']

    # Check structure
    print(f"  Rows: {ws.max_row}")
    print(f"  Columns: {ws.max_column}")

    # Print all headers
    print("\n  Row 1 (Month Headers):")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            print(f"    Col {col}: {cell.value}")

    print("\n  Row 2 (Field Headers):")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        if cell.value:
            print(f"    Col {col}: {cell.value}")

    print("\n  Row 3 (Data):")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col)
        if cell.value:
            print(f"    Col {col}: {cell.value}")

    # Verify merged cells
    merged_ranges = list(ws.merged_cells.ranges)
    print(f"\n  Merged cell ranges: {len(merged_ranges)}")
    for mr in merged_ranges:
        print(f"    - {mr}")

    # Final validation
    assert ws.max_row == 3, f"Expected 3 rows, got {ws.max_row}"
    assert ws.max_column == 8, f"Expected 8 columns, got {ws.max_column}"
    assert len(merged_ranges) == 5, f"Expected 5 merged ranges, got {len(merged_ranges)}"

    print("\n✓ Excel file is VALID (no corruption)")
    print(f"✓ File saved at: {tmp_path}")
    print("\nYou can manually open this file in Excel to verify:")
    print(f"  open {tmp_path}")

    return tmp_path


if __name__ == "__main__":
    try:
        path = test_realistic_bench_allocation()
        print("\n" + "=" * 70)
        print("SUCCESS: Excel file is valid and can be opened")
        print("=" * 70)
    except Exception as e:
        print("\n" + "=" * 70)
        print("FAILURE: Excel file has issues")
        print("=" * 70)
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
