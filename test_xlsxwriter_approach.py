"""
Test using xlsxwriter instead of openpyxl.
xlsxwriter is often more compatible with Excel's strict format requirements.
"""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from io import BytesIO
import pandas as pd
import tempfile


def test_xlsxwriter_approach():
    """Create Excel with xlsxwriter engine."""
    print("\n" + "=" * 70)
    print("TESTING XLSXWRITER APPROACH")
    print("=" * 70)

    # Sample data
    data = {
        'Main LOB': ['Amisys Medicaid DOMESTIC', 'Facets Commercial DOMESTIC'],
        'State': ['TX', 'CA'],
        'Case Type': ['Claims', 'Enrollment'],
        'Case ID': ['CL-001', 'EN-002'],
        'Jun-25 Client Forecast': ['1000', None],
        'Jun-25 FTE Required': ['20', None],
        'Jun-25 FTE Available': ['25 (20)', None],
        'Jun-25 Capacity': ['1125 (1000)', None],
        'Jul-25 Client Forecast': [None, '2000'],
        'Jul-25 FTE Required': [None, '40'],
        'Jul-25 FTE Available': [None, '50 (40)'],
        'Jul-25 Capacity': [None, '2500 (2000)'],
    }

    df = pd.DataFrame(data)

    print("\n1. Creating Excel with xlsxwriter...")
    excel_buffer = BytesIO()

    # Use xlsxwriter engine
    with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
        # Write data without headers (we'll add custom headers)
        df.to_excel(writer, sheet_name='Changes', index=False, header=False, startrow=2)

        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Changes']

        # Define formats
        header_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#366092',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })

        field_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#5B9BD5',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })

        # Static columns
        static_columns = ['Main LOB', 'State', 'Case Type', 'Case ID']

        # Write static column headers (merged vertically)
        for idx, col_name in enumerate(static_columns):
            worksheet.merge_range(0, idx, 1, idx, col_name, header_format)

        # Month headers and field headers
        months = ['Jun-25', 'Jul-25']
        fields = ['Client Forecast', 'FTE Required', 'FTE Available', 'Capacity']

        col_idx = len(static_columns)
        for month in months:
            # Merge month header horizontally across 4 columns
            start_col = col_idx
            end_col = col_idx + len(fields) - 1
            worksheet.merge_range(0, start_col, 0, end_col, month, header_format)

            # Write field headers in row 2
            for field in fields:
                worksheet.write(1, col_idx, field, field_format)
                col_idx += 1

        # Set column widths
        for idx, col_name in enumerate(df.columns):
            max_len = max(df[col_name].astype(str).map(len).max(), len(col_name))
            worksheet.set_column(idx, idx, min(max_len + 2, 50))

    print("   ✓ Excel created")

    # Save to file
    excel_buffer.seek(0)
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        tmp.write(excel_buffer.read())
        tmp_path = tmp.name

    print(f"\n2. Saved to: {tmp_path}")

    # Verify with openpyxl
    print("\n3. Verifying with openpyxl...")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(tmp_path)
        ws = wb['Changes']
        print(f"   ✓ openpyxl can open: {ws.max_row} rows, {ws.max_column} cols")

        # Check for None values in headers
        none_count_r1 = sum(1 for col in range(1, ws.max_column + 1)
                           if ws.cell(row=1, column=col).value is None)
        none_count_r2 = sum(1 for col in range(1, ws.max_column + 1)
                           if ws.cell(row=2, column=col).value is None)

        print(f"   Row 1 None values: {none_count_r1}")
        print(f"   Row 2 None values: {none_count_r2}")

    except Exception as e:
        print(f"   ✗ openpyxl failed: {e}")

    print("\n" + "=" * 70)
    print("✓ XLSXWRITER TEST COMPLETE")
    print("=" * 70)
    print(f"\nFile: {tmp_path}")
    print("\nPlease open this file in Excel and check if repair is needed.")
    print("If this works without repair, we should switch to xlsxwriter.")

    return tmp_path


if __name__ == "__main__":
    try:
        # Check if xlsxwriter is installed
        import xlsxwriter
        print("✓ xlsxwriter is installed")
    except ImportError:
        print("✗ xlsxwriter not installed")
        print("Install with: pip install xlsxwriter")
        exit(1)

    try:
        path = test_xlsxwriter_approach()
        print(f"\n✓ Test completed: {path}")
    except Exception as e:
        print(f"\n✗ Test failed: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
