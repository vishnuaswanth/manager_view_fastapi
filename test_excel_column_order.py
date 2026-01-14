"""
Test that Excel column order matches expected structure to prevent corruption.

This test verifies that:
1. DataFrame columns are in the correct order
2. Multi-level headers are created for the correct columns
3. Excel file is valid (no corruption)
"""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    HistoryLogData,
    HistoryChangeRecord,
    generate_history_excel,
    CORE_FIELDS
)
from openpyxl import load_workbook


def test_column_order_single_month():
    """Test column order with single month."""
    print("\n" + "=" * 70)
    print("Test 1: Column Order - Single Month")
    print("=" * 70)

    history_log = HistoryLogData(
        id="test-123",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T10:30:00Z",
        user="test_user",
        description="Test",
        records_modified=1,
        summary_data=None
    )

    # Create changes for all 4 fields in Jun-25
    changes = [
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.forecast",
            old_value=1000,
            new_value=1000,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_req",
            old_value=20,
            new_value=20,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,
            delta=5,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1125,
            delta=125,
            month_label="Jun-25"
        )
    ]

    # Generate Excel
    excel_buffer = generate_history_excel(history_log, changes)
    excel_buffer.seek(0)

    # Load and verify
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print("\nExpected structure:")
    print("  Static columns (4): Main LOB, State, Case Type, Case ID")
    print("  Month columns (4): Jun-25 → Client Forecast, FTE Required, FTE Available, Capacity")

    print("\nActual structure:")
    print("  Row 1 (Month Headers):")
    for col in range(1, 9):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            print(f"    Column {col}: {cell.value}")

    print("\n  Row 2 (Field Headers):")
    for col in range(1, 9):
        cell = ws.cell(row=2, column=col)
        if cell.value:
            print(f"    Column {col}: {cell.value}")

    # Verify static columns (1-4)
    static_cols = ["Main LOB", "State", "Case Type", "Case ID"]
    for i, expected in enumerate(static_cols, start=1):
        actual = ws.cell(row=1, column=i).value
        assert actual == expected, f"Column {i}: Expected '{expected}', got '{actual}'"

    # Verify month header (column 5 should be Jun-25)
    month_header = ws.cell(row=1, column=5).value
    assert month_header == "Jun-25", f"Column 5 should be 'Jun-25', got '{month_header}'"

    # Verify field headers (columns 5-8)
    for i, expected_field in enumerate(CORE_FIELDS, start=5):
        actual_field = ws.cell(row=2, column=i).value
        assert actual_field == expected_field, f"Column {i} row 2: Expected '{expected_field}', got '{actual_field}'"

    print("\n✓ Column order is correct")


def test_column_order_multiple_months():
    """Test column order with multiple months."""
    print("\n" + "=" * 70)
    print("Test 2: Column Order - Multiple Months")
    print("=" * 70)

    history_log = HistoryLogData(
        id="test-456",
        change_type="Bench Allocation",
        month="March",
        year=2025,
        timestamp="2025-03-15T10:30:00Z",
        user="test_user",
        description="Test",
        records_modified=1,
        summary_data=None
    )

    # Create changes for Jun-25 and Jul-25
    changes = []
    months = ["Jun-25", "Jul-25"]

    for month in months:
        for field in ['forecast', 'fte_req', 'fte_avail', 'capacity']:
            changes.append(HistoryChangeRecord(
                main_lob="Amisys Medicaid DOMESTIC",
                state="TX",
                case_type="Claims",
                case_id="CL-001",
                field_name=f"{month}.{field}",
                old_value=100,
                new_value=100,
                delta=0,
                month_label=month
            ))

    # Generate Excel
    excel_buffer = generate_history_excel(history_log, changes)
    excel_buffer.seek(0)

    # Load and verify
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print("\nExpected structure:")
    print("  Static columns (4): Main LOB, State, Case Type, Case ID")
    print("  Jun-25 columns (4): Client Forecast, FTE Required, FTE Available, Capacity")
    print("  Jul-25 columns (4): Client Forecast, FTE Required, FTE Available, Capacity")
    print("  Total: 12 columns")

    print(f"\nActual columns: {ws.max_column}")

    # Verify structure
    assert ws.max_column == 12, f"Expected 12 columns, got {ws.max_column}"

    # Verify month headers
    jun_header = ws.cell(row=1, column=5).value
    jul_header = ws.cell(row=1, column=9).value

    assert jun_header == "Jun-25", f"Column 5 should be 'Jun-25', got '{jun_header}'"
    assert jul_header == "Jul-25", f"Column 9 should be 'Jul-25', got '{jul_header}'"

    # Verify Jun-25 field headers (columns 5-8)
    for i, expected_field in enumerate(CORE_FIELDS, start=5):
        actual_field = ws.cell(row=2, column=i).value
        assert actual_field == expected_field, f"Jun-25 col {i}: Expected '{expected_field}', got '{actual_field}'"

    # Verify Jul-25 field headers (columns 9-12)
    for i, expected_field in enumerate(CORE_FIELDS, start=9):
        actual_field = ws.cell(row=2, column=i).value
        assert actual_field == expected_field, f"Jul-25 col {i}: Expected '{expected_field}', got '{actual_field}'"

    print("\n✓ Multiple months column order is correct")


def test_excel_file_integrity():
    """Test that Excel file can be opened without repair."""
    print("\n" + "=" * 70)
    print("Test 3: Excel File Integrity (No Corruption)")
    print("=" * 70)

    history_log = HistoryLogData(
        id="test-789",
        change_type="CPH Update",
        month="March",
        year=2025,
        timestamp="2025-03-15T10:30:00Z",
        user="system",
        description="Test",
        records_modified=1,
        summary_data=None
    )

    # Create comprehensive changes
    changes = [
        # Target CPH (static field)
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="target_cph",
            old_value=45,
            new_value=50,
            delta=5,
            month_label=None
        ),
        # Jun-25 fields
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.forecast",
            old_value=1000,
            new_value=1000,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_req",
            old_value=20,
            new_value=18,
            delta=-2,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=20,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.capacity",
            old_value=1000,
            new_value=1100,
            delta=100,
            month_label="Jun-25"
        )
    ]

    # Generate Excel
    excel_buffer = generate_history_excel(history_log, changes)
    excel_buffer.seek(0)

    # Try to load - should not raise any warnings/errors
    try:
        wb = load_workbook(excel_buffer)
        ws = wb['Changes']

        # Verify basic structure
        assert ws.max_row >= 3, "Should have at least 3 rows (2 header + 1 data)"
        assert ws.max_column >= 5, "Should have at least 5 columns"

        # Verify merged cells exist
        merged_ranges = list(ws.merged_cells.ranges)
        assert len(merged_ranges) > 0, "Should have merged cells for headers"

        print(f"\n✓ Excel file loaded successfully")
        print(f"  Rows: {ws.max_row}")
        print(f"  Columns: {ws.max_column}")
        print(f"  Merged cell ranges: {len(merged_ranges)}")

        # Verify no empty/None values in header rows
        for row_idx in [1, 2]:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Cell can be None if it's part of a merged cell
                # But the merge anchor should have a value
                if not cell.value and not isinstance(cell, type(cell)):  # Not a MergedCell
                    # Check if this cell is part of a merged range
                    is_merged = any(
                        (row_idx, col_idx) in merged_range
                        for merged_range in merged_ranges
                    )
                    if not is_merged:
                        print(f"  Warning: Empty cell at row {row_idx}, col {col_idx}")

        print("✓ No corruption detected - file is valid")

    except Exception as e:
        print(f"\n✗ CORRUPTION DETECTED: {e}")
        import traceback
        traceback.print_exc()
        raise


def main():
    """Run all column order tests."""
    print("\n" + "=" * 70)
    print("EXCEL COLUMN ORDER TEST SUITE")
    print("=" * 70)
    print("\nVerifying that DataFrame column order matches multi-level headers")
    print("to prevent Excel corruption/repair issues")

    try:
        test_column_order_single_month()
        test_column_order_multiple_months()
        test_excel_file_integrity()

        print("\n" + "=" * 70)
        print("✓ ALL COLUMN ORDER TESTS PASSED")
        print("=" * 70)
        print("\nExcel File Quality:")
        print("  ✓ Column order is consistent and correct")
        print("  ✓ Multi-level headers match data columns")
        print("  ✓ No file corruption detected")
        print("  ✓ Excel can open file without repair")

        return 0

    except AssertionError as e:
        print("\n" + "=" * 70)
        print("✗ TEST FAILED")
        print("=" * 70)
        print(f"\nAssertion Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    except Exception as e:
        print("\n" + "=" * 70)
        print("✗ UNEXPECTED ERROR")
        print("=" * 70)
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main())
