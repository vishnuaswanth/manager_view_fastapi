"""
Trace exactly what happens during header creation to find the corruption source.
"""

import sys
import os
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, project_root)

from code.logics.history_excel_generator import (
    HistoryLogData,
    HistoryChangeRecord,
    _prepare_pivot_data,
    CORE_FIELDS
)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import pandas as pd


def trace_excel_creation():
    """Trace step-by-step what happens during Excel creation."""
    print("\n" + "=" * 70)
    print("TRACING EXCEL CREATION STEP-BY-STEP")
    print("=" * 70)

    # Create test data
    changes = [
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.forecast",
            old_value=1000,
            new_value=1000,
            delta=0,
            month_label="Jun-25"
        ),
        HistoryChangeRecord(
            main_lob="Amisys Medicaid DOMESTIC",
            state="TX",
            case_type="Claims",
            case_id="CL-001",
            field_name="Jun-25.fte_avail",
            old_value=20,
            new_value=25,
            delta=5,
            month_label="Jun-25"
        )
    ]

    # Get pivot data
    pivot_data, month_labels, static_columns = _prepare_pivot_data(changes)

    print("\n1. Pivot Data Structure:")
    print(f"   Months: {month_labels}")
    print(f"   Static columns: {static_columns}")
    print(f"   Pivot data keys: {list(pivot_data[0].keys()) if pivot_data else []}")

    # Build column order
    column_order = list(static_columns)
    for month_label in month_labels:
        for field in CORE_FIELDS:
            column_order.append(f"{month_label} {field}")

    print(f"\n2. Column Order: {column_order}")

    # Create DataFrame
    df = pd.DataFrame(pivot_data, columns=column_order)
    print(f"\n3. DataFrame columns: {list(df.columns)}")

    # Write to Excel with pandas
    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Changes', index=False)

    # Check what pandas wrote
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb['Changes']

    print(f"\n4. AFTER pandas.to_excel():")
    print(f"   Total rows: {ws.max_row}")
    print(f"   Total columns: {ws.max_column}")
    print("\n   Row 1 (pandas header):")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=1, column=col_idx)
        print(f"     Col {col_idx}: '{cell.value}'")

    print("\n   Row 2 (pandas data):")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=2, column=col_idx)
        print(f"     Col {col_idx}: '{cell.value}'")

    # Now insert row (simulating what _create_multilevel_headers does)
    print("\n5. AFTER ws.insert_rows(1):")
    ws.insert_rows(1)
    print(f"   Total rows: {ws.max_row}")

    print("\n   Row 1 (newly inserted, empty):")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=1, column=col_idx)
        print(f"     Col {col_idx}: '{cell.value}' (type: {type(cell.value).__name__})")

    print("\n   Row 2 (old pandas header, shifted down):")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=2, column=col_idx)
        print(f"     Col {col_idx}: '{cell.value}'")

    print("\n   Row 3 (old pandas data, shifted down):")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=3, column=col_idx)
        print(f"     Col {col_idx}: '{cell.value}'")

    # Now write our custom headers
    print("\n6. Writing custom multi-level headers...")

    col_idx = 1

    # Static columns
    for col_name in static_columns:
        ws.cell(row=1, column=col_idx, value=col_name)
        ws.cell(row=2, column=col_idx, value="")  # Overwrite old pandas header
        col_idx += 1

    # Month headers
    for month_label in month_labels:
        start_col = col_idx
        ws.cell(row=1, column=start_col, value=month_label)

        # Empty strings for merged cells in row 1
        for i in range(start_col + 1, start_col + len(CORE_FIELDS)):
            ws.cell(row=1, column=i, value="")

        # Field headers in row 2
        for field_name in CORE_FIELDS:
            ws.cell(row=2, column=col_idx, value=field_name)
            col_idx += 1

    print("\n7. AFTER writing custom headers (before merge):")
    print("\n   Row 1:")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=1, column=col_idx)
        print(f"     Col {col_idx}: '{cell.value}'")

    print("\n   Row 2:")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=2, column=col_idx)
        print(f"     Col {col_idx}: '{cell.value}'")

    # Apply merges
    print("\n8. Applying merges...")
    merge_ranges = []

    # Merge static columns vertically
    for i in range(1, len(static_columns) + 1):
        col_letter = get_column_letter(i)
        merge_range = f'{col_letter}1:{col_letter}2'
        merge_ranges.append(merge_range)
        ws.merge_cells(merge_range)
        print(f"   Merged: {merge_range}")

    # Merge month headers horizontally
    start_col = len(static_columns) + 1
    for month_label in month_labels:
        end_col = start_col + len(CORE_FIELDS) - 1
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        merge_range = f'{start_letter}1:{end_letter}1'
        merge_ranges.append(merge_range)
        ws.merge_cells(merge_range)
        print(f"   Merged: {merge_range}")
        start_col = end_col + 1

    print(f"\n9. AFTER merging (final state):")
    print(f"   Total merged ranges: {len(merge_ranges)}")

    print("\n   Row 1:")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=1, column=col_idx)
        cell_type = type(cell).__name__
        print(f"     Col {col_idx}: '{cell.value}' (type: {cell_type})")

    print("\n   Row 2:")
    for col_idx in range(1, min(ws.max_column + 1, 10)):
        cell = ws.cell(row=2, column=col_idx)
        cell_type = type(cell).__name__
        print(f"     Col {col_idx}: '{cell.value}' (type: {cell_type})")

    # Save and check if Excel can open it
    import tempfile
    with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    print(f"\n10. File saved at: {tmp_path}")
    print("    Try opening in Excel to see if repair is needed.")

    return tmp_path


if __name__ == "__main__":
    trace_excel_creation()
